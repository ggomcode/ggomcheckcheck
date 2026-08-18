import io
import os
import re
import json
import traceback
import requests
import pandas as pd
import numpy as np
import streamlit as st
import streamlit.components.v1 as stc
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase.cidfonts import UnicodeCIDFont

# ==============================================================================
# 페이지 기본 설정 & CSS 커스텀 스타일링 (Light Theme & Premium Card Aesthetics)
# ==============================================================================
def get_favicon_b64(icon_path: str = "favicon.png") -> str:
    import base64
    if os.path.exists(icon_path):
        try:
            with open(icon_path, "rb") as f:
                return base64.b64encode(f.read()).decode("utf-8")
        except Exception:
            pass
    return ""


st.set_page_config(
    page_title="꼼체크체크 - 생기부 AI 정밀 검증 시스템",
    page_icon="favicon.png" if os.path.exists("favicon.png") else "📋",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for Modern, Premium UI with Balanced, Elegant Borders
st.markdown("""
    <style>
    .block-container {
        padding-top: 1.8rem;
        padding-bottom: 3rem;
    }
    .main-header {
        background: linear-gradient(135deg, #0F172A 0%, #1E293B 100%);
        padding: 1.8rem 2rem;
        border-radius: 14px;
        color: #F8FAFC;
        margin-bottom: 1.8rem;
        box-shadow: 0 4px 20px rgba(0, 0, 0, 0.08);
        border: 1px solid #334155;
    }
    .main-header h1 {
        color: #38BDF8;
        font-size: 1.85rem;
        font-weight: 700;
        margin-bottom: 0.4rem;
    }
    .main-header p {
        color: #94A3B8;
        font-size: 0.95rem;
        margin: 0;
    }
    .metric-card {
        background-color: #FFFFFF;
        border: 1px solid #CBD5E1;
        border-radius: 10px;
        padding: 1.2rem;
        text-align: center;
        box-shadow: 0 2px 6px rgba(0,0,0,0.03);
    }
    .metric-card .val {
        font-size: 1.6rem;
        font-weight: 700;
        color: #0F172A;
    }
    .metric-card .lbl {
        font-size: 0.85rem;
        color: #64748B;
        font-weight: 500;
        margin-top: 0.2rem;
    }
    .student-card {
        background: #F8FAFC;
        border: 1px solid #E2E8F0;
        border-left: 4px solid #0EA5E9;
        border-radius: 8px;
        padding: 1.2rem 1.4rem;
        margin-bottom: 1rem;
    }
    .subject-badge {
        display: inline-block;
        background-color: #0EA5E9;
        color: white;
        font-weight: 600;
        font-size: 0.8rem;
        padding: 0.25rem 0.6rem;
        border-radius: 6px;
        margin-bottom: 0.6rem;
    }
    .char-badge {
        float: right;
        font-size: 0.8rem;
        color: #64748B;
        font-weight: 500;
    }
    /* Hide file uploader size limit and format text (200MB per file • XLSX, XLS) */
    [data-testid="stFileUploaderInstructions"],
    [data-testid="stFileUploaderInstructions"] *,
    [data-testid="stFileUploaderDropzoneInstructions"],
    [data-testid="stFileUploaderDropzoneInstructions"] *,
    div[data-testid="stFileUploader"] small,
    section[data-testid="stFileUploader"] small,
    div[data-testid="stFileUploader"] small *,
    section[data-testid="stFileUploader"] small * {
        display: none !important;
        visibility: hidden !important;
        height: 0 !important;
        width: 0 !important;
        font-size: 0 !important;
        margin: 0 !important;
        padding: 0 !important;
    }

    /* Hide Sidebar Collapse Button & Header */
    [data-testid="stSidebarCollapseButton"],
    [data-testid="stSidebarHeader"],
    [data-testid="stSidebarHeader"] button,
    section[data-testid="stSidebar"] button[aria-label*="close"],
    section[data-testid="stSidebar"] button[aria-label*="Collapse"],
    section[data-testid="stSidebar"] button[aria-label*="접기"],
    [data-testid="stSidebarNav"],
    #MainMenu,
    header[data-testid="stHeader"],
    footer,
    [data-testid="stToolbar"],
    .stAppDeployButton,
    div[data-testid="stAppDeployButton"],
    div[data-testid="stDecoration"] {
        display: none !important;
        visibility: hidden !important;
        height: 0 !important;
        width: 0 !important;
        margin: 0 !important;
        padding: 0 !important;
    }

    section[data-testid="stSidebar"] * {
        color: #0F172A !important;
    }
    section[data-testid="stSidebar"] .stCaption, 
    section[data-testid="stSidebar"] [data-testid="stCaptionContainer"] {
        color: #64748B !important;
        font-weight: 500 !important;
    }

    /* Force Expander Header & Labels onto 1 Single Line without Arrow Icon */
    [data-testid="stExpander"] details summary svg,
    .stExpander summary svg,
    details summary svg {
        display: none !important;
        visibility: hidden !important;
        width: 0 !important;
        height: 0 !important;
    }
    [data-testid="stExpander"] details summary,
    .stExpander summary {
        padding-left: 0.75rem !important;
    }
    [data-testid="stExpander"] details summary p,
    [data-testid="stExpander"] details summary span,
    .stExpander summary p,
    .stExpander summary span {
        white-space: nowrap !important;
        overflow: visible !important;
        text-overflow: clip !important;
        font-size: 0.95rem !important;
    }
    /* Sub-expander headers (e.g. Gemini API Key 발급 가이드) font size reduced by 1pt */
    section[data-testid="stSidebar"] [data-testid="stExpander"] [data-testid="stExpander"] details summary p,
    section[data-testid="stSidebar"] [data-testid="stExpander"] [data-testid="stExpander"] details summary span {
        font-size: 0.88rem !important;
        font-weight: 600 !important;
    }

    /* Sidebar Expander Color Themes & Balanced Borders */
    section[data-testid="stSidebar"] [data-testid="stExpander"] {
        border-radius: 12px !important;
        overflow: hidden !important;
        margin-bottom: 0.85rem !important;
        box-shadow: 0 2px 6px rgba(0, 0, 0, 0.03) !important;
    }
    section[data-testid="stSidebar"] [data-testid="stExpander"] details summary {
        padding: 0.7rem 0.9rem !important;
        border-radius: 10px !important;
        transition: all 0.2s ease-in-out !important;
    }

    /* 1. AI (LLM) API 설정 (Indigo Blue Theme) */
    section[data-testid="stSidebar"] [data-testid="stExpander"]:nth-of-type(1) {
        border: 1px solid #A5B4FC !important;
        background-color: #FFFFFF !important;
    }
    section[data-testid="stSidebar"] [data-testid="stExpander"]:nth-of-type(1) details summary {
        background: linear-gradient(135deg, #EEF2FF 0%, #E0E7FF 100%) !important;
        border-bottom: 1px solid #C7D2FE !important;
    }
    section[data-testid="stSidebar"] [data-testid="stExpander"]:nth-of-type(1) details summary,
    section[data-testid="stSidebar"] [data-testid="stExpander"]:nth-of-type(1) details summary * {
        color: #3730A3 !important;
        font-weight: 700 !important;
        font-size: 0.95rem !important;
    }

    /* 2. 엑셀 파일 업로드 (Mint / Emerald Green Theme) */
    section[data-testid="stSidebar"] [data-testid="stExpander"]:nth-of-type(2) {
        border: 1px solid #6EE7B7 !important;
        background-color: #FFFFFF !important;
    }
    section[data-testid="stSidebar"] [data-testid="stExpander"]:nth-of-type(2) details summary {
        background: linear-gradient(135deg, #ECFDF5 0%, #D1FAE5 100%) !important;
        border-bottom: 1px solid #A7F3D0 !important;
    }
    section[data-testid="stSidebar"] [data-testid="stExpander"]:nth-of-type(2) details summary,
    section[data-testid="stSidebar"] [data-testid="stExpander"]:nth-of-type(2) details summary * {
        color: #065F46 !important;
        font-weight: 700 !important;
        font-size: 0.95rem !important;
    }

    /* 3. 데이터 관리 메뉴 (Lavender Purple Theme) */
    section[data-testid="stSidebar"] [data-testid="stExpander"]:nth-of-type(3) {
        border: 1px solid #C4B5FD !important;
        background-color: #FFFFFF !important;
    }
    section[data-testid="stSidebar"] [data-testid="stExpander"]:nth-of-type(3) details summary {
        background: linear-gradient(135deg, #F5F3FF 0%, #EDE9FE 100%) !important;
        border-bottom: 1px solid #DDD6FE !important;
    }
    section[data-testid="stSidebar"] [data-testid="stExpander"]:nth-of-type(3) details summary,
    section[data-testid="stSidebar"] [data-testid="stExpander"]:nth-of-type(3) details summary * {
        color: #5B21B6 !important;
        font-weight: 700 !important;
        font-size: 0.95rem !important;
    }

    /* API Key Label - Bold Dark Red */
    section[data-testid="stSidebar"] div[data-testid="stTextInput"] label,
    section[data-testid="stSidebar"] div[data-testid="stTextInput"] label p {
        color: #B91C1C !important;
        font-weight: 700 !important;
    }

    /* Form Controls - Clear Distinct Background (#E2E8F0) & Sleek Slate Border */
    div[data-baseweb="select"],
    div[data-baseweb="select"] > div,
    div[data-baseweb="input"],
    div[data-baseweb="input"] > div,
    div[data-baseweb="base-input"],
    div[data-testid="stSelectbox"] [data-baseweb="select"] > div,
    div[data-testid="stTextInput"] [data-baseweb="input"] > div,
    div:has(> input[role="combobox"]),
    div:has(> input[type="password"]),
    div:has(> input[aria-label*="Key"]),
    div:has(> input.ef5iutx2),
    div:has(> input.e1fp86qc1) {
        background-color: #E2E8F0 !important;
        border: 1.5px solid #94A3B8 !important;
        border-radius: 8px !important;
        box-shadow: 0 1px 3px rgba(15, 23, 42, 0.08) !important;
    }

    /* Input Text & Selected Value Color */
    div[data-baseweb="select"] * {
        color: #0F172A !important;
        font-weight: 600 !important;
    }
    div[data-baseweb="input"] input,
    input[role="combobox"],
    input[type="password"],
    input[type="text"],
    input {
        color: #0F172A !important;
        font-weight: 600 !important;
        background-color: transparent !important;
    }

    div[data-baseweb="select"] svg,
    div[data-baseweb="input"] svg {
        fill: #1E293B !important;
        color: #1E293B !important;
    }

    div[data-baseweb="select"] > div:hover,
    div[data-baseweb="input"] > div:hover {
        border-color: #2563EB !important;
        background-color: #DBEAFE !important;
    }

    /* File Uploader - Minimal spacing */
    [data-testid="stFileUploader"],
    .stFileUploader {
        padding: 0 !important;
        border: none !important;
        background: transparent !important;
    }

    /* General Buttons - Balanced Clean Border */
    div.stButton > button:not([kind="primary"]):not([data-testid="stBaseButton-primary"]) {
        background-color: #FFFFFF !important;
        background: #FFFFFF !important;
        color: #0F172A !important;
        border: 1px solid #94A3B8 !important;
        border-radius: 8px !important;
        font-weight: 600 !important;
        transition: all 0.2s ease !important;
        box-shadow: 0 1px 3px rgba(0,0,0,0.04) !important;
    }
    div.stButton > button:not([kind="primary"]):not([data-testid="stBaseButton-primary"]) * {
        color: #0F172A !important;
    }

    /* Download Buttons - Clean Blue Border */
    [data-testid="stDownloadButton"] button,
    [data-testid="stDownloadButton"] a {
        background-color: #EFF6FF !important;
        color: #1D4ED8 !important;
        border: 1px solid #3B82F6 !important;
        border-radius: 8px !important;
        font-weight: 700 !important;
    }
    [data-testid="stDownloadButton"] button *,
    [data-testid="stDownloadButton"] a * {
        color: #1D4ED8 !important;
    }

    /* Primary Button */
    button[kind="primary"],
    [data-testid="stBaseButton-primary"] {
        background: linear-gradient(135deg, #2563EB 0%, #1D4ED8 100%) !important;
        color: #FFFFFF !important;
        border: 1px solid #1D4ED8 !important;
        border-radius: 8px !important;
        font-weight: 700 !important;
        box-shadow: 0 4px 12px rgba(37, 99, 235, 0.25) !important;
    }
    button[kind="primary"] *,
    [data-testid="stBaseButton-primary"] * {
        color: #FFFFFF !important;
    }
    </style>
""", unsafe_allow_html=True)


# ==============================================================================
# 1. 지침 MD 파일 동적 로더
# ==============================================================================
def load_guideline_content() -> str:
    """
    data/ 폴더 내의 '학교생활기록부_기재_및_검증_지침.md', '세부 규칙.md' 지침을 실시간으로 읽어옵니다.
    분석 실행 시 항상 최신 파일 내용을 반영합니다.
    """
    contents = []
    
    # 1. 기본 기재 및 검증 지침 파일
    base_candidates = [
        "data/학교생활기록부_기재_및_검증_지침.md", 
        "data/학교생활기록부_기재_및_검증_지침.MD",
        "data/기재요령.md"
    ]
    for bp in base_candidates:
        if os.path.exists(bp):
            try:
                with open(bp, "r", encoding="utf-8") as f:
                    contents.append(f.read().strip())
                break
            except Exception as e:
                pass

    # 2. 사용자 세부 규칙 파일 (data/세부 규칙.md 등 - 최우선 적용 지침)
    detail_candidates = [
        "data/세부 규칙.md",
        "data/세부_규칙.md",
        "data/세부규칙.md"
    ]
    for dp in detail_candidates:
        if os.path.exists(dp):
            try:
                with open(dp, "r", encoding="utf-8") as f:
                    detail_text = f.read().strip()
                    if detail_text:
                        contents.append(f"\n\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n【사용자 정의 세부 규칙 (최우선 적용 지침)】\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n{detail_text}")
                break
            except Exception as e:
                pass

    return "\n\n".join(contents)


# ==============================================================================
# 2. LLM AI 분석 엔진 (Gemini / OpenAI / Claude API 파이프라인)
# ==============================================================================
def fetch_available_gemini_models(api_key: str):
    """
    Google AI Studio API를 통해 사용자 키에서 지원 가능한 Gemini 모델 목록을 동적으로 전수 조회합니다.
    pageSize=100 및 페이지네이션을 지원하여 모든 최신 모델을 누락 없이 가져옵니다.
    """
    if not api_key or len(api_key.strip()) < 10:
        return []
    try:
        clean_key = api_key.strip()
        models = []
        page_token = None
        
        while True:
            url = f"https://generativelanguage.googleapis.com/v1beta/models?pageSize=100&key={clean_key}"
            if page_token:
                url += f"&pageToken={page_token}"
            
            res = requests.get(url, timeout=8)
            if res.status_code != 200:
                break
            
            data = res.json()
            raw_models = data.get("models", [])
            for m in raw_models:
                methods = m.get("supportedGenerationMethods", [])
                if "generateContent" in methods:
                    m_id = m.get("name", "").replace("models/", "")
                    # 오직 순수 Gemini 모델만 필터링 (gemma, learnlm, bison 등 비-Gemini 모델 및 임베딩 제외)
                    if not m_id.lower().startswith("gemini"):
                        continue
                    if any(skip in m_id.lower() for skip in ["embedding", "aqa", "imagen", "whisper", "tts"]):
                        continue
                    
                    display = m.get("displayName", m_id)
                    
                    score = 50
                    if "2.5-flash" in m_id.lower():
                        score = 110
                    elif "3.1-pro" in m_id.lower():
                        score = 105
                    elif "1.5-pro-002" in m_id.lower():
                        score = 100
                    elif "1.5-pro" in m_id.lower():
                        score = 95
                    elif "2.0-flash" in m_id.lower():
                        score = 90
                    elif "1.5-flash" in m_id.lower():
                        score = 85
                    elif "2.0-pro" in m_id.lower():
                        score = 80
                    elif "pro" in m_id.lower():
                        score = 75
                    elif "flash" in m_id.lower():
                        score = 70

                    models.append({
                        "id": m_id,
                        "display": f"{display} ({m_id})",
                        "score": score
                    })
            
            page_token = data.get("nextPageToken")
            if not page_token:
                break

        # 중복 제거 및 점수 순 정렬
        seen_ids = set()
        unique_models = []
        for m in sorted(models, key=lambda x: x["score"], reverse=True):
            if m["id"] not in seen_ids:
                seen_ids.add(m["id"])
                unique_models.append(m)

        print(f"[DEBUG models] Fetched {len(unique_models)} dynamic models from Google API: {[m['id'] for m in unique_models]}")
        return unique_models
    except Exception as e:
        print(f"[DEBUG models] Error fetching models: {e}")
        pass
    return []


def clean_json_response(raw_text: str) -> str:
    """
    AI 응답에서 마크다운 코드블록(```json ... ```)을 정제합니다.
    """
    if not raw_text:
        return ""
    clean = raw_text.strip()
    if clean.startswith("```"):
        clean = re.sub(r'^```[a-zA-Z]*\n?', '', clean)
        clean = re.sub(r'```$', '', clean)
        clean = clean.strip()
    return clean


def robust_json_parse(raw_text: str):
    """
    AI 응답을 다양한 손상/미완결/마크다운 형식에서도 결함 없이 JSON 데이터로 복원합니다.
    """
    if not raw_text:
        return None
    clean = clean_json_response(raw_text)
    
    # 1. 표준 json.loads 시도
    try:
        return json.loads(clean)
    except Exception:
        pass

    # 2. 가장 바깥쪽 { ... } 또는 [ ... ] 추출
    m = re.search(r'(\{[\s\S]*\}|\[[\s\S]*\])', clean)
    if m:
        try:
            return json.loads(m.group(1))
        except Exception:
            pass

    # 3. 미완결 잘림 자동 완성 시도 (누락된 닫는 괄호 보완)
    for close_suffix in ['\n]}', '\n}', ']}', '}']:
        try:
            return json.loads(clean + close_suffix)
        except Exception:
            pass

    # 4. 정규식을 통한 개별 검출 객체 단위 전수 추출 (가장 강력한 Fallback)
    item_pattern = re.compile(r'\{[^{}]*(?:student_id|original_text|수정전)[^{}]*\}', re.DOTALL)
    extracted_items = []
    for match in item_pattern.finditer(clean):
        raw_item = match.group(0)
        try:
            extracted_items.append(json.loads(raw_item))
        except Exception:
            try:
                fixed_item = re.sub(r',\s*\}', '}', raw_item)
                extracted_items.append(json.loads(fixed_item))
            except Exception:
                pass
    if extracted_items:
        return {"results": extracted_items}

    return None


def call_llm_api_for_audit(provider: str, api_key: str, model_name: str, records_data: list, guideline_text: str, progress_callback=None) -> list:
    """
    생기부 기록 데이터를 적정 배치(Batch)로 나누어 API 쿼터(250k 토큰 제한)를 초과하지 않도록 분할 호출합니다.
    429 Rate Limit/503 Service Unavailable 오류 발생 시 자동 지연 재시도(Exponential Retry)를 수행합니다.
    """
    import time
    if not api_key:
        raise ValueError("API Key가 설정되지 않았습니다. 사이드바에서 AI API Key를 입력해 주세요.")

    import threading
    from concurrent.futures import ThreadPoolExecutor, as_completed

    # 스마트 배치 엔진 - Rate Limit 친화적 설계
    # Gemini 무료 티어: 15 RPM → 최소 4초 간격 필요
    # 배치 크기를 키워 API 호출 수를 최소화하되, 정확도가 유지되는 범위 내에서 조절
    from collections import defaultdict
    student_grouped = defaultdict(list)
    for r in records_data:
        st_key = (safe_str(r.get("학번", "")), safe_str(r.get("이름", "")))
        student_grouped[st_key].append(r)

    batches = []
    target_chars_budget = 12000   # 12,000자 (A4 약 4~5장) - 호출 수 최소화
    max_items_per_chunk = 30      # 과목 30개까지 한 배치

    for st_key, recs in student_grouped.items():
        total_chars = sum(len(x.get("기록텍스트", "")) for x in recs)
        
        if total_chars <= target_chars_budget and len(recs) <= max_items_per_chunk:
            # 학생 1명 통째 (창체/행특/일반 세특)
            batches.append(recs)
        else:
            # 대용량: 이수학년별 분할
            grade_sub_grouped = defaultdict(list)
            for r in recs:
                g_key = safe_str(r.get("이수학년", ""))
                grade_sub_grouped[g_key].append(r)
            
            for g_key, g_recs in grade_sub_grouped.items():
                g_chars = sum(len(x.get("기록텍스트", "")) for x in g_recs)
                if g_chars <= target_chars_budget and len(g_recs) <= max_items_per_chunk:
                    batches.append(g_recs)
                else:
                    for i in range(0, len(g_recs), max_items_per_chunk):
                        batches.append(g_recs[i:i + max_items_per_chunk])

    if not batches and records_data:
        batches = [records_data]

    total_batches = len(batches)
    all_results = []
    completed_batches = 0
    print(f"[INFO] Total batches to process: {total_batches} (students: {len(student_grouped)})")

    prompt_instructions = f"""
당신은 대한민국 교육부 학교생활기록부(창체, 세특, 행특) 오탈자 교정 및 기재지침 검증 최고 권위의 감사 전문가 AI입니다.
전달된 학생의 모든 활동 및 과목별 서술문을 글자 하나, 단어 하나 단위로 전수 정밀 감사하여, 발견된 모든 오탈자, 맞춤법 오류, 띄어쓰기 오류, 기재금지어를 빠짐없이 'results' 목록에 담아 반환하십시오.

[참고 지침 문서]
{guideline_text}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
【검증 핵심 대상 및 등급(severity) 기준】
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
1. [수정 필수] - 법적 지침 위반, 맞춤법/오탈자/띄어쓰기/조사/특수문자 오류 (가장 엄격하고 적극적으로 검출할 것):
   - 조사 호응 및 조사 선택 오류 (받침 유무 및 문맥 호응):
     * 받침 없는 체언 뒤 조사 '을' 오용: '카드뉴스을' ➔ '카드뉴스를', '생산 인구을' ➔ '생산 인구를', '갯수을' ➔ '개수를'
     * 문맥상 잘못된 조사 사용: '목민심서를 나타난' ➔ '목민심서에 나타난', '역사적 배경으로 기반으로' ➔ '역사적 배경을 기반으로'
     * 조사 오탈자 / 불필요한 끝조사: '활동체 참여' ➔ '활동에 참여', '명극즉 과찰이다의' ➔ '명즉즉 과찰이다'
   - 자판 오타, 철자 오탈자, 음절 오기입, 단어 중복:
     * 자판 오타 / 자음모음 연타: '주의 집중력읖' ➔ '주의 집중력을', '장ㄹ주행' ➔ '자율주행', '학업 성취르르' ➔ '학업 성취를', '방았음' ➔ '받았음', '빗방물' ➔ '빗방울', '찾기위애' ➔ '찾기 위해'
     * 음절/철자 오탈자: '합습' ➔ '학습', '편근' ➔ '편견', '실생황' ➔ '실생활', '색체' ➔ '색채', '해결해기' ➔ '해결하기', '구체화려' ➔ '구체화하려', '제작하하는데' ➔ '제작하는 데', '젠더 간 갈들을' ➔ '젠더 간 갈등을'
     * 단어 중복 기재: '유통 유통망' ➔ '유통망'
   - 맞춤법 및 피동/사동/어미/사이시옷 오류:
     * '높힘.' ➔ '높임.', '신뢰도를 높혔음.' ➔ '신뢰도를 높였음.'
     * '영상이였으며' ➔ '영상이었으며'
     * '갯수' ➔ '개수'
     * '띔' ➔ '띰', '안되' ➔ '안 돼', '되서' ➔ '돼서', '치뤄' ➔ '치러', '바램' ➔ '바람', '만듬' ➔ '만듦', '도우는' ➔ '돕는', '맞추다'/'맞히다', '낳다'/'낫다'
   - 문맥상 부자연스러운 단어 오용 / 동음이의어 오탈자:
     * '정서함.' ➔ '작성함.' (문맥상 글이나 문장을 지어 기록한다는 의미의 오탈자)
     * '캐슬을' ➔ '캔슬을' (캔슬 컬처 오타)
     * '정병 대상 인구수' ➔ '병역 대상 인구수' (군 복무 대상 인구)
     * '제풀이' ➔ '풀이'
   - 불필요하거나 잘못 삽입된 특수문자/공백/가운뎃점:
     * '기사를·분석해' ➔ '기사를 분석해' (단어/조사 사이에 불필요하게 삽입된 가운뎃점(·) 제거)
     * '생성형 , 인공지능의' ➔ '생성형 인공지능의' (쉼표 앞 불필요한 공백 제거)
     * 따옴표('"), 쉼표(,), 마침표(.), 느낌표(!), 물음표(?), 콜론(:), 괄호 외 불필요 특수기호(★, ◆, ~, @, #, $ 등) 제거
   - 띄어쓰기 규정 위반:
     * '할수있다' ➔ '할 수 있다', '찾기위해' ➔ '찾기 위해', '초등학교때' ➔ '초등학교 때', '배운점' ➔ '배운 점', '느낀점' ➔ '느낀 점', '수업시간' ➔ '수업 시간', '다양한활동' ➔ '다양한 활동' 등 모든 한국어 띄어쓰기 규정 위반
   - 법적 기재 금지어 및 미허용 브랜드명/약어:
     * 특정 상호나 브랜드명 노출('프OOOO, 봄O' 등) ➔ 익명화 처리 또는 일반 명사로 기재
     * 입력 불가 약어/용어: 'OTT' ➔ '동영상 플랫폼' 등 매핑표 지정 대체어로 교정 제안
     * 공인어학성적, 교외 수상, 논문/학회지/도서출판/특허, 사교육/학원명, 영재교육원, 부모 직업, 특정 대학명/기관명 등
   - 창체 영역별 이수시간 0시간 (순회/위탁 등 특수 사유 서술문 학생 제외)

2. [수정 권장] - 문맥 및 서술 완성도 개선, 표준 외래어 표기:
   - 표준 외래어 표기법 및 권장 용어:
     * '프리젠테이션' ➔ '프레젠테이션'
     * '어플' ➔ '앱'
   - 문맥에 맞는 적절한 어휘 및 띄어쓰기 개선:
     * '외성적' ➔ '외형적'
     * '서본결' ➔ '서결론'
     * '냉동 선생산' ➔ '냉동 선행 생산'
   - 주어-서술어 불일치, 지나치게 어색한 비문 또는 문장 구조 다듬기

3. [검토 권장] - 서술 관점 개선 및 소극적 서술 완화:
   - 학생의 주관적 추측/독백 서술 ➔ 교사 관찰 중심 전환:
     * '~기대해 봄.', '~느낌.', '~배움.', '~다짐함.' ➔ '~에 대한 관심이 높음.', '~모습을 보임.'
   - 소극적 서술 완화:
     * '주제탐구 활동에 참여하려고 노력함.' ➔ '주제 탐구 활동에 적극적으로 참여함.' / '주제 탐구 활동에 참여함.'
   - 보다 학술적이고 객관적인 용어로 표현 개선:
     * '인공지능 그림' ➔ '생성형 인공지능 활용 이미지'

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
【예외 및 주의사항 (오탐 방지)】
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
- 도서명(책 제목) 및 저자명은 지침 2.1.4에 따라 전 영역 기재 가능하므로 절대 오류로 지적하지 마십시오.
- PPT, IT, AI, DNA, RNA, STEAM, SW, VR, AR, UCC, SNS, POPS, PD, TV, CEO 등 표준 영문 약어는 학교생활기록부 기재요령상 허용된 합법적 표기입니다.
- '(1학기)', '(2학기)' 학기 표기 및 동아리활동 '(동아리명)(이수시간)' 표기는 나이스(NEIS) 표준 서식이므로 절대 오류로 지적하지 마십시오.
- 'E사', 'A사', 'B사' 등 알파벳 1글자 기업 블라인드 표기는 정당한 익명화 방식입니다.
- 날짜/기간 표기('2025.07.08.' 등)는 지양 사항일 뿐 오류로 지적하지 마십시오.
- 엑셀 페이지 나눔으로 단어가 끊긴 경우(예: '비교 분석하고', '신뢰감 있는') 쪼개진 글자를 임의로 삭제하거나 없는 단어를 환각으로 지어내지 마십시오.
- 'reason(수정해야하는 이유나 근거)' 항목을 작성할 때는 조항 번호(예: '지침 2.1.4' 등) 없이 맞춤법, 띄어쓰기, 어미 교정 등 구체적인 이유만 작성하십시오.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
【출력 형식 (JSON Schema)】
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
반드시 오직 아래의 JSON 데이터만 반환하십시오. 마크다운 추가 텍스트 없이 순수한 JSON만 반환해야 합니다:
{{
  "results": [
    {{
      "student_id": "학번 (5자리 숫자로 작성, 예: 30101)",
      "student_name": "학생 이름",
      "category": "구분 (창체 / 세특 / 행발 중 하나)",
      "taken_grade": "이수학년 (예: 1학년, 2학년, 3학년)",
      "sub_category": "세부 (창체는 '자율활동', '동아리활동', '진로활동' 중 1개, 세특은 구체적 과목명 또는 '개세특', 행특은 '행동특성')",
      "original_text": "오류가 발견된 수정 전 단어/문구",
      "suggested_text": "올바르게 교정된 수정 후 추천 문구",
      "reason": "수정해야 하는 명확한 이유나 근거 (조항 번호 없이 구체적 이유 기술)",
      "severity": "수정 필수, 수정 권장, 검토 권장 중 하나 (맞춤법/오탈자/띄어쓰기/금지어는 반드시 '수정 필수')"
    }}
  ]
}}
"""

    try:
        from streamlit.runtime.scriptrunner import add_script_run_ctx, get_script_run_ctx
        cur_ctx = get_script_run_ctx()
    except Exception:
        cur_ctx = None

    # Rate Limit 친화적 순차 처리 (Gemini 무료 티어: 15 RPM)
    import threading
    _api_lock = threading.Lock()
    _last_call_time = [0.0]  # mutable for closure
    MIN_CALL_INTERVAL = 4.5  # 초 (15 RPM = 4초 간격 + 여유 0.5초)

    def rate_limited_api_call(b_idx, batch_data):
        """Rate limit을 준수하며 API를 호출하는 함수"""
        st_num = safe_str(batch_data[0].get("학번", "")) if batch_data else ""
        st_name = safe_str(batch_data[0].get("이름", "")) if batch_data else ""
        grades = list(set(safe_str(x.get("이수학년", "")) for x in batch_data if x.get("이수학년")))
        grade_tag = f" [{grades[0]}]" if len(grades) == 1 else ""
        student_label = f"{st_num} {st_name}{grade_tag}".strip() if (st_num or st_name) else f"{b_idx+1}번 배치"

        mini_payload = []
        for r in batch_data:
            mini_payload.append({
                "student_id": safe_str(r.get("학번", "")),
                "student_name": safe_str(r.get("이름", "")),
                "category": safe_str(r.get("구분", "")),
                "taken_grade": safe_str(r.get("이수학년", "")),
                "sub_category": safe_str(r.get("세부", "")),
                "text": safe_str(r.get("기록텍스트", ""))
            })
        data_payload_text = json.dumps(mini_payload, ensure_ascii=False, indent=2)
        full_prompt = f"{prompt_instructions}\n\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n【검증할 학생 서술문 데이터】\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n{data_payload_text}"
        print(f"[DEBUG batch#{b_idx}] student={student_label}, items={len(mini_payload)}, prompt_len={len(full_prompt)}")

        session = requests.Session()
        session.headers.update({
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        })

        max_retries = 8
        success = False
        raw_response_text = ""

        for attempt in range(max_retries):
            # Rate Limit 준수: 마지막 호출로부터 최소 MIN_CALL_INTERVAL초 대기
            with _api_lock:
                elapsed = time.time() - _last_call_time[0]
                if elapsed < MIN_CALL_INTERVAL:
                    wait = MIN_CALL_INTERVAL - elapsed
                    print(f"[DEBUG batch#{b_idx}] Rate limit wait: {wait:.1f}s")
                    time.sleep(wait)
                _last_call_time[0] = time.time()

            try:
                if provider.lower() == "gemini":
                    base_model = model_name if model_name else "gemini-2.0-flash"
                    # 503/429 발생 시 다른 안정적 엔드포인트로 자동 폴백 전환
                    if attempt >= 2:
                        fallback_list = ["gemini-2.0-flash", "gemini-1.5-flash", "gemini-2.5-flash", "gemini-flash-latest"]
                        current_model = fallback_list[(attempt - 2) % len(fallback_list)]
                        print(f"[DEBUG batch#{b_idx}] 503/429 Fallback: Switching model to '{current_model}'")
                    else:
                        current_model = base_model

                    url = f"https://generativelanguage.googleapis.com/v1beta/models/{current_model}:generateContent?key={api_key}"
                    payload = {
                        "contents": [{"parts": [{"text": full_prompt}]}],
                        "generationConfig": {
                            "responseMimeType": "application/json",
                            "temperature": 0.1,
                            "maxOutputTokens": 8192
                        }
                    }
                    res = session.post(url, json=payload, timeout=120)
                    print(f"[DEBUG batch#{b_idx}] Gemini ({current_model}) HTTP {res.status_code} (attempt {attempt+1})")
                    if res.status_code == 404:
                        # 404(모델 지원 종료/미존재) 발생 시 즉시 유효한 최신 모델로 전환
                        print(f"[DEBUG batch#{b_idx}] Model '{current_model}' 404 NOT_FOUND. Fallback to alternative model.")
                        fallback_list = ["gemini-3.1-pro-preview", "gemini-1.5-pro", "gemini-2.0-flash", "gemini-1.5-flash"]
                        current_model = fallback_list[attempt % len(fallback_list)]
                        time.sleep(1.0)
                        continue
                    if res.status_code == 429:
                        # 429 전용 장기 대기: 15초, 30초, 45초, 60초...
                        backoff = min(15 * (attempt + 1), 90)
                        print(f"[DEBUG batch#{b_idx}] 429 Rate Limit! Backing off {backoff}s...")
                        time.sleep(backoff)
                        continue
                    if res.status_code in [500, 502, 503, 504]:
                        backoff = (attempt + 1) * 5
                        print(f"[DEBUG batch#{b_idx}] Server error {res.status_code}, backing off {backoff}s...")
                        time.sleep(backoff)
                        continue
                    if res.status_code != 200:
                        print(f"[DEBUG batch#{b_idx}] ERROR: {res.text[:500]}")
                        raise RuntimeError(f"Gemini API 호출 실패 ({res.status_code}): {res.text}")
                    res_json = res.json()
                    raw_response_text = res_json['candidates'][0]['content']['parts'][0]['text']
                    print(f"[DEBUG batch#{b_idx}] SUCCESS! response_len={len(raw_response_text)}, preview={raw_response_text[:150]}")
                    success = True
                    break

                elif provider.lower() == "openai":
                    model = model_name if model_name else "gpt-4o-mini"
                    url = "https://api.openai.com/v1/chat/completions"
                    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
                    payload = {
                        "model": model,
                        "messages": [
                            {"role": "system", "content": "You are a helpful assistant that outputs student record verification results in JSON schema."},
                            {"role": "user", "content": full_prompt}
                        ],
                        "response_format": {"type": "json_object"}
                    }
                    res = session.post(url, headers=headers, json=payload, timeout=120)
                    if res.status_code == 429:
                        backoff = min(15 * (attempt + 1), 90)
                        time.sleep(backoff)
                        continue
                    if res.status_code in [500, 502, 503, 504]:
                        time.sleep((attempt + 1) * 5)
                        continue
                    if res.status_code != 200:
                        raise RuntimeError(f"OpenAI API 호출 실패 ({res.status_code}): {res.text}")
                    res_json = res.json()
                    raw_response_text = res_json['choices'][0]['message']['content']
                    success = True
                    break

                elif provider.lower() == "claude":
                    url = "https://api.anthropic.com/v1/messages"
                    headers = {"x-api-key": api_key, "anthropic-version": "2023-06-01", "Content-Type": "application/json"}
                    payload = {
                        "model": "claude-3-5-haiku-20241022",
                        "max_tokens": 4000,
                        "system": "Return strictly a JSON object with 'results' array as requested.",
                        "messages": [{"role": "user", "content": full_prompt}]
                    }
                    res = session.post(url, headers=headers, json=payload, timeout=120)
                    if res.status_code == 429:
                        backoff = min(15 * (attempt + 1), 90)
                        time.sleep(backoff)
                        continue
                    if res.status_code in [500, 502, 503, 504]:
                        time.sleep((attempt + 1) * 5)
                        continue
                    if res.status_code != 200:
                        raise RuntimeError(f"Claude API 호출 실패 ({res.status_code}): {res.text}")
                    res_json = res.json()
                    raw_response_text = res_json['content'][0]['text']
                    success = True
                    break

            except Exception as req_err:
                if attempt == max_retries - 1:
                    print(f"[DEBUG batch#{b_idx}] FINAL FAILURE after {max_retries} attempts: {req_err}")
                    raise req_err
                time.sleep((attempt + 1) * 5)

        batch_findings = []
        if success and raw_response_text:
            try:
                parsed = robust_json_parse(raw_response_text)
                if parsed is None:
                    print(f"[DEBUG batch#{b_idx}] parsed is None! clean_json preview: {clean_json_response(raw_response_text)[:300]}")
                elif isinstance(parsed, list):
                    batch_findings = parsed
                    print(f"[DEBUG batch#{b_idx}] parsed as list, findings={len(batch_findings)}")
                elif isinstance(parsed, dict):
                    print(f"[DEBUG batch#{b_idx}] parsed as dict, keys={list(parsed.keys())}")
                    if "results" in parsed and isinstance(parsed["results"], list):
                        batch_findings = parsed["results"]
                    elif "findings" in parsed and isinstance(parsed["findings"], list):
                        batch_findings = parsed["findings"]
                    elif "errors" in parsed and isinstance(parsed["errors"], list):
                        batch_findings = parsed["errors"]
                    elif "data" in parsed and isinstance(parsed["data"], list):
                        batch_findings = parsed["data"]
                    else:
                        for v in parsed.values():
                            if isinstance(v, list) and len(v) > 0 and isinstance(v[0], dict):
                                batch_findings = v
                                break
                    print(f"[DEBUG batch#{b_idx}] final findings={len(batch_findings)}")
            except Exception as outer_e:
                print(f"[DEBUG batch#{b_idx}] OUTER EXCEPTION: {outer_e}")
        else:
            print(f"[DEBUG batch#{b_idx}] SKIPPED: success={success}, raw_len={len(raw_response_text)}")

        return (b_idx, student_label, batch_findings)

    # 순차 처리 (Rate Limit 준수) - 1개씩 순서대로 호출하되 API 간격을 자동 조절
    for idx, b in enumerate(batches):
        if progress_callback:
            st_num = safe_str(b[0].get("학번", "")) if b else ""
            st_name = safe_str(b[0].get("이름", "")) if b else ""
            grades = list(set(safe_str(x.get("이수학년", "")) for x in b if x.get("이수학년")))
            grade_tag = f" [{grades[0]}]" if len(grades) == 1 else ""
            lbl = f"{st_num} {st_name}{grade_tag}".strip()
            progress_callback("start", idx, total_batches, lbl)

        b_idx, student_label, batch_findings = rate_limited_api_call(idx, b)
        completed_batches += 1
        if batch_findings:
            all_results.extend(batch_findings)
        if progress_callback:
            progress_callback("finish", completed_batches, total_batches, student_label, batch_findings)

    if 'data_store' in st.session_state and 'hope_blank_records' in st.session_state['data_store']:
        for hb in st.session_state['data_store']['hope_blank_records']:
            all_results.append(hb)

    filtered_results = []
    for res_item in all_results:
        orig = str(res_item.get('original_text', '')).strip()
        sugg = str(res_item.get('suggested_text', '')).strip()
        reason = str(res_item.get('reason', '')).strip()

        # 세특 세부 컬럼 정규화: '세부능력및특기사항' -> '개세특'
        if str(res_item.get('category', '')) == '세특':
            if str(res_item.get('sub_category', '')).strip() in ['세부능력및특기사항', '세부능력 및 특기사항', '세특', '세부', 'nan', '']:
                res_item['sub_category'] = '개세특'

        # 1. Skip single letter company initial (e.g. 'E사', 'A사', 'B사')
        if re.match(r'^[A-Za-z]사$', orig):
            continue

        # 2. Skip if original_text and suggested_text are identical (no correction needed)
        if orig == sugg:
            continue

        # 3. 이수시간 관련 점검은 오직 '0시간'인 경우만 검출 (사용자 지침: 이수시간 오류는 0시간일때만 체크해)
        if '이수시간' in orig or '이수시간' in sugg or '이수시간' in reason or re.match(r'^\d+\s*시간$', orig) or (orig.isdigit() and orig != '0'):
            if '0시간' not in orig and '0시간' not in reason and orig != '0':
                continue

        # 4. Skip if original_text is plain '희망분야' (without 공란 tag)
        if orig in ['희망분야', '희망 분야']:
            continue

        # 6. Skip false errors on Dongari default format (동아리명)(이수시간)
        sub_c = str(res_item.get('sub_category', '')).strip()
        if '동아리' in sub_c or re.match(r'^\([가-힣a-zA-Z0-9\s·/]+\)\s*\(\d+시간\)', orig):
            if ('동아리' in reason or '괄호' in reason or '활동 명칭' in reason) and ('제거' in reason or '삭제' in reason or '표기' in reason or '명시' in reason or '수정' in reason):
                continue
            if re.match(r'^\([가-힣a-zA-Z0-9\s·/]+\)\s*\(\d+시간\)', orig):
                continue

        # 7. Skip false errors on page-split fragment syllables (림., 고, 함, 시, 활 등)
        if any(kw in reason for kw in ['종결되지', '도중에 끊', '어미가 완성', '문맥을 보완', '내용을 보완', '미완성', '완결되지', '따옴표로 끝', '데이터 분할', '데이터 병합', '잘려 나간']):
            continue

        if orig in ['림.', '고', '함', '시', '활', '림', '알', '함.'] and any(kw in reason for kw in ['잘려', '분할', '유실', '불완전']):
            continue

        # 8. Skip allowed English acronyms (PPT, IT, AI, DNA, RNA, STEAM, SW, VR, AR, UCC, SNS, POPS, PD, TV, CEO 등)
        allowed_english = ['PPT', 'IT', 'AI', 'DNA', 'RNA', 'STEAM', 'SW', 'VR', 'AR', 'UCC', 'SNS', 'POPS', 'PD', 'TV', 'CEO', 'MBTI', '앱', 'APP', 'PDF']
        if orig.upper() in allowed_english or any(ae in orig.upper() for ae in ['PPT', 'IT', 'AI', 'SW', 'VR', 'AR', 'UCC', 'SNS', 'POPS', 'PD', 'TV', 'CEO']):
            if any(kw in reason for kw in ['영문', '외국어', '알파벳', '한글', '대체', '사용 지양']):
                continue

        # 9. Skip date and period expressions (2025.07.08. 등 날짜 표기는 지양 사항일 뿐 오류 아님)
        if '날짜' in reason or '기간' in reason or '일자' in reason or re.search(r'\d{4}\.\d{2}\.\d{2}', orig) or re.search(r'\d{1,2}월\s*\d{1,2}일', orig):
            continue

        # 10. Skip internal sentence emotion/reflection modifiers ('느꼈던', '생각한', '깨달은' 등)
        if any(em in orig for em in ['느꼈던', '생각한', '깨달은', '고민한', '배운']) and any(kw in reason for kw in ['내면', '심리', '학생 입장', '서술 지양', '활동 중심']):
            continue

        # 13. Skip false error on '소셜벤처' being mistaken for ESG
        if ('소셜벤처' in orig or '소셜 벤처' in orig) and any(kw in reason for kw in ['ESG', '약어', '대체', '지속가능']):
            continue

        # 14. Skip false errors on fragmented word hallucination ('과학적 이', '알', '함', '시' 등)
        if any(kw in reason for kw in ['중간에 잘려', '단어가 잘려', '완성된 문구', '유실된 단어', '단어를 문맥에 맞게 복구', '잘린 어미', '어미를 보완']):
            continue

        filtered_results.append(res_item)

    return filtered_results

    return filtered_results


# ==============================================================================
# 3. 텍스트 정제 및 지능형 텍스트 결합 헬퍼 함수
# ==============================================================================
def safe_val(val, default=""):
    """
    pd.Series, numpy array, list, 또는 단일 스칼라 값을 안전하게 단일 값으로 변환합니다.
    """
    if val is None:
        return default
    if isinstance(val, (pd.Series, np.ndarray, list)):
        if len(val) == 0:
            return default
        val = val.iloc[0] if isinstance(val, pd.Series) else val[0]
    if pd.isna(val):
        return default
    return val


def safe_str(val, default="") -> str:
    """
    pd.Series, numpy array 등이 들어와도 Ambiguous Error 없이 안전하게 단일 문자열로 변환합니다.
    """
    v = safe_val(val, default)
    return str(v).strip()


def safe_notna(val) -> bool:
    """
    pd.Series, numpy array 등이 들어와도 Ambiguous Error 없이 안전하게 pd.notna 검사를 수행합니다.
    """
    if val is None:
        return False
    if isinstance(val, (pd.Series, np.ndarray, list)):
        if len(val) == 0:
            return False
        val = val.iloc[0] if isinstance(val, pd.Series) else val[0]
    return bool(pd.notna(val))


def safe_isna(val) -> bool:
    return not safe_notna(val)


def clean_text_content(text) -> str:
    if safe_isna(text):
        return ""
    text_str = safe_str(text)
    # 줄바꿈(\n\n) 구조는 개세특 분리를 위해 보존하면서 탭과 불필요한 공백만 정제
    text_str = text_str.replace('\r\n', '\n').replace('\r', '\n')
    text_str = re.sub(r'[ \t]+', ' ', text_str)
    text_str = re.sub(r'\n{3,}', '\n\n', text_str)
    return text_str.strip()


def get_changche_area_from_cell(area_cell, text_cell) -> str:
    """
    엑셀 셀에 입력된 영역명 컬럼 값 및 기재 텍스트 형태(예: 접두어/키워드)를
    직접 분석하여 '자율활동', '동아리활동', '진로활동' 중 올바른 영역 1개를 추출합니다.
    """
    a_str = safe_str(area_cell)
    t_str = safe_str(text_cell)

    if '자율' in a_str:
        return '자율활동'
    if '동아리' in a_str:
        return '동아리활동'
    if '진로' in a_str:
        return '진로활동'

    if '동아리' in t_str or re.match(r'^\([가-힣a-zA-Z0-9\s·/]+\)\s*\(\d+시간\)', t_str):
        return '동아리활동'
    if '진로' in t_str:
        return '진로활동'
    if '자율' in t_str:
        return '자율활동'

    return '자율활동'


def check_jinro_hope_field(row_dict: dict, row_series: pd.Series = None) -> tuple:
    """
    진로활동 영역에서 I열 특기사항이 '희망사항' 항목일 때 M열(학생 희망진로) 내용 검사.
    Returns: (is_hope_row: bool, hope_value: str)
    """
    row_values = [safe_str(v) for v in row_dict.values() if safe_notna(v)]
    for idx, val in enumerate(row_values):
        if len(val) <= 20 and ('희망사항' in val or '희망분야' in val or '희망 분야' in val):
            hope_val = ""
            if row_series is not None and len(row_series) > 12:
                candidate_m = safe_str(row_series.iloc[12])
                if candidate_m and candidate_m.lower() not in ['nan', 'none', '희망분야', '희망 분야', '희망사항', '특기사항']:
                    hope_val = candidate_m
            if not hope_val and idx + 1 < len(row_values):
                candidate = row_values[idx + 1].strip()
                if candidate and candidate.lower() not in ['nan', 'none', '희망분야', '희망 분야', '희망사항', '특기사항']:
                    hope_val = candidate
            return True, hope_val
    return False, ""


def is_meaningful_concatenation(base_token: str, append_token: str) -> bool:
    """
    페이지 나눔선을 기준으로 앞 페이지 끝 단어(음절)와 뒷 페이지 맨 앞 단어(음절)를 붙였을 때,
    그 문장 내에서 유의미한 하나의 단어, 형태소(조사, 어미, 접사 결합) 또는 복합어가 되는지 판단합니다.
    """
    b = base_token.strip()
    a = append_token.strip()
    if not b or not a:
        return False

    # 1. 조사 / 어미 / 접미사 패턴 검사
    # 예: 어머니 + 가, 포트폴리오 + 를, 추진 + 함., 수행 + 하였음, 바탕 + 으로, 알 + 림., 선정 + 고, 노력 + 함
    josa_eomi_pattern = (
        r'^(이|가|을|를|은|는|의|에|에서|에게|한테|과|와|으로|로|도|만|까지|부터|마저|조차|'
        r'(이)?며|(이)?나|(이)?든|함|음|슴|했음|였음|하며|하여|하고|한|할|된|됨|되어|되고|되는|된다|'
        r'적|성|율|률|력|화|스럽|스러운|스럽게|시키|시킴|시켰|시작|바탕|림|림\.|들|들이|들을|들의|들에게)'
    )
    if re.match(josa_eomi_pattern, a):
        return True

    # 2. 잘려나간 용언 어미 / 시제 / 연결어미 접미 (예: 하 + 였음 -> 하였음, 나타 + 냄 -> 나타냄, 이끌 + 어냄)
    if re.match(r'^(였음|았음|었음|였으며|았으며|었으며|였고|았고|었고|여|아|어|면서|으면서|도록|고자|려고|으려고|는지|은지|ㄴ지|ㄹ수록|을수록|거나|든지|더미|거리|살이|냄|김)', a):
        return True

    # 3. 1~2음절 한자어/고유어 형태소 합성 검증 사전 (빈출 생기부 음절 결합)
    morpheme_pairs = [
        ('제', '작'), ('제', '출'), ('제', '시'), ('제', '공'), ('제', '안'),
        ('설', '명'), ('설', '계'), ('설', '정'), ('설', '립'),
        ('탐', '구'), ('탐', '색'), ('탐', '독'),
        ('발', '표'), ('발', '견'), ('발', '전'), ('발', '달'),
        ('분', '석'), ('분', '류'), ('분', '야'),
        ('이', '해'), ('이', '유'), ('이', '론'), ('이', '용'),
        ('능', '력'), ('역', '할'), ('역', '사'), ('역', '량'),
        ('활', '용'), ('활', '동'), ('활', '성'),
        ('해', '결'), ('해', '석'), ('해', '당'),
        ('학', '습'), ('학', '교'), ('학', '문'),
        ('수', '행'), ('수', '학'), ('수', '업'),
        ('참', '여'), ('참', '고'), ('참', '가'),
        ('토', '론'), ('토', '의'),
        ('어머', '니'), ('알', '고리즘'), ('알', '게'), ('알', '림'),
        ('정', '리'), ('실', '천'), ('모', '습'), ('과', '정'),
        ('성', '취'), ('노', '력'), ('열', '정'), ('태', '도'),
        ('시', '각'), ('시', '도'), ('시', '작'),
        ('기', '반'), ('기', '초'), ('기', '능')
    ]
    for prefix, suffix in morpheme_pairs:
        if b.endswith(prefix) and a.startswith(suffix):
            return True

    # 4. 앞 토큰이 1음절로 잘린 불완전 형태소인 경우 (예: '알' + '고리즘', '제' + '작을', '탐' + '구를')
    if len(b) == 1 and re.match(r'^[가-힣]', b) and re.match(r'^[가-힣]', a):
        return True

    return False


def smart_concatenate_text(base_text: str, append_text: str) -> str:
    base_clean = base_text.strip()
    append_clean = append_text.strip()

    if not base_clean:
        return append_clean
    if not append_clean:
        return base_clean

    last_char = base_clean[-1]
    first_char = append_clean[0]

    # 1. 문장 부호 및 괄호 뒤 연결 시 띄어쓰기 적용
    if last_char in ['.', '!', '?', ':', ';', ',', ')', ']', '}', '"', "'"]:
        return f"{base_clean} {append_clean}"

    # 2. 앞 페이지 마지막 단어와 뒷 페이지 첫 단어 추출
    base_tokens = base_clean.split()
    append_tokens = append_clean.split()

    last_token = base_tokens[-1] if base_tokens else ""
    first_token = append_tokens[0] if append_tokens else ""

    # 따옴표나 백틱 등 특수문자 제거 후 순수 어절 비교
    clean_last_token = re.sub(r'[\'\"`]', '', last_token)
    clean_first_token = re.sub(r'[\'\"`]', '', first_token)

    # 3. 페이지 나눔 기준 단어/음절 결합 유의미성 판단 (Meaningful Word Concatenation)
    if is_meaningful_concatenation(clean_last_token, clean_first_token):
        return base_clean + append_clean

    # 4. 앞 단어가 이미 조사가 붙었거나 완성된 용언 활용형인 경우 -> 띄어쓰기 유지
    # 예: '발표하고' + '토론함' -> '발표하고 토론함', '보고서를' + '제출함' -> '보고서를 제출함'
    complete_ending_pattern = r'(하고|하며|하여|했으나|있고|보이며|같음|으로|에서|에게|을|를|이|가|은|는|과|와)$'
    if re.search(complete_ending_pattern, clean_last_token):
        return f"{base_clean} {append_clean}"

    # 5. 둘 다 한글 완성형 단어인 경우 기본 띄어쓰기 적용
    if re.match(r'[가-힣a-zA-Z0-9]', last_char) and re.match(r'[가-힣a-zA-Z0-9]', first_char):
        return f"{base_clean} {append_clean}"

    return f"{base_clean} {append_clean}"


# ==============================================================================
# 4. 동적 컬럼 자동 매핑 및 가비지 컬럼 제거 헬퍼
# ==============================================================================
def is_header_or_footer_row(row_dict: dict, num_col=None, name_col=None) -> bool:
    vals = [safe_str(v) for v in row_dict.values() if safe_notna(v)]
    combined = "".join(vals).replace(" ", "")

    num_val_str = safe_str(row_dict.get(num_col, '')).replace(" ", "") if num_col else ""
    name_val_str = safe_str(row_dict.get(name_col, '')).replace(" ", "") if name_col else ""

    if re.search(r'<[가-힣\sㆍ·/]+>', num_val_str):
        return True

    if '번호' in num_val_str or '성명' in name_val_str or '번 호' in num_val_str or '성 명' in name_val_str:
        return True

    if len(combined) < 40 and ('영역' in combined or '창의적체험활동' in combined or '세부능력' in combined or '행동특성' in combined):
        if '시간' in combined or '특기사항' in combined or '상황' in combined or '의견' in combined:
            return True

    # NEIS 엑셀 페이지 꼬리말 구조 정밀 검사: '/' 셀이 있고 페이지 숫자 및 학교명이 존재하는 행
    if '/' in vals and (any('고등학교' in v or '학교' in v for v in vals) or any(re.match(r'^\d+\.?\d*$', v) for v in vals)):
        return True

    if len(combined) < 50 and ('학교' in combined or '사용자명' in combined or '페이지' in combined):
        if any(re.search(r'^\d+$', v.strip()) for v in vals) or '/' in combined:
            return True

    if len(combined) < 40 and re.search(r'^\s*\d+/\d+\.?\d*\s*$', combined):
        return True

    return False


def deduplicate_columns(df: pd.DataFrame) -> pd.DataFrame:
    cols = list(df.columns)
    seen = {}
    new_cols = []
    for c in cols:
        c_str = str(c).strip()
        if c_str in seen:
            seen[c_str] += 1
            new_cols.append(f"{c_str}_{seen[c_str]}")
        else:
            seen[c_str] = 0
            new_cols.append(c_str)
    df.columns = new_cols
    return df


def get_safe_series(df: pd.DataFrame, col_name: str) -> pd.Series:
    if col_name not in df.columns:
        return pd.Series([""] * len(df), index=df.index)
    val = df[col_name]
    if isinstance(val, pd.DataFrame):
        return val.iloc[:, 0]
    return val


def clean_display_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """
    NEIS 셀 병합으로 인해 생겨난 nan, nan_1~nan_7, Unnamed, None 등
    불필요한 가비지 컬럼을 깔끔하게 제거하고 1부터 시작하는 인덱스를 할당합니다.
    """
    if df is None or df.empty:
        return df
    clean_df = df.copy()
    drop_cols = [
        c for c in clean_df.columns 
        if str(c).startswith('_') 
        or str(c).strip().lower().startswith('nan') 
        or str(c).strip().lower().startswith('unnamed')
        or str(c).strip() == 'None'
    ]
    if drop_cols:
        clean_df = clean_df.drop(columns=drop_cols, errors='ignore')
    clean_df.index = range(1, len(clean_df) + 1)
    return clean_df


def detect_columns(df: pd.DataFrame) -> tuple:
    df_processed = deduplicate_columns(df.copy())

    header_idx = None
    for idx in range(min(20, len(df_processed))):
        row_vals = [safe_str(v).replace(" ", "") for v in df_processed.iloc[idx].values if safe_notna(v)]
        row_str = "".join(row_vals)
        if ('번호' in row_str or '번' in row_str) and ('성명' in row_str or '이름' in row_str):
            header_idx = idx
            break

    if header_idx is not None:
        h1 = [safe_str(v) for v in df_processed.iloc[header_idx].values]
        skip_rows = 1
        
        if header_idx + 1 < len(df_processed):
            next_row = df_processed.iloc[header_idx + 1]
            col_b_val = safe_str(next_row.iloc[1]).strip() if len(next_row) > 1 else ""
            is_student_row = bool(re.match(r'^\d+$', col_b_val))
            
            if not is_student_row:
                h2_str = "".join([safe_str(v).replace(" ", "") for v in next_row.values if safe_notna(v)])
                if any(k in h2_str for k in ['영역', '특기사항', '시간', '구분', '내용']):
                    h2 = [safe_str(v) for v in next_row.values]
                    combined = []
                    for c1, c2 in zip(h1, h2):
                        if c2 and c2.lower() != 'nan' and c2.lower() != 'none':
                            combined.append(c2)
                        elif c1 and c1.lower() != 'nan' and c1.lower() != 'none':
                            combined.append(c1)
                        else:
                            combined.append('nan')
                    h1 = combined
                    skip_rows = 2

        df_processed.columns = h1
        df_processed = df_processed.iloc[header_idx + skip_rows:].reset_index(drop=True)

    df_processed = deduplicate_columns(df_processed)

    columns = list(df_processed.columns)
    mapped = {
        'num_col': None,
        'name_col': None,
        'content_col': None,
        'area_col': None,
        'grade_col': None,
        'extra_cols': []
    }

    num_keywords = ['번호', '학생번호', '순번', 'no', 'num', 'id', '학번']
    name_keywords = ['성명', '이름', '학생명', '성 명', 'name', '학생']
    content_keywords = [
        '세부능력 및 특기사항', '세부능력및특기사항', '행동특성 및 종합의견', '행동특성및종합의견', '행동특성',
        '창의적 체험활동 영역별 특기사항', '창체', '특기사항', '기록 내용', '기록내용', '내용', '종합의견', '세특'
    ]
    area_keywords = ['영역', '활동영역', '구분', '과목', '과목명']
    grade_keywords = ['학년', '학 년']

    for col in columns:
        col_clean = str(col).strip().replace(" ", "").lower()
        if not mapped['num_col']:
            for kw in num_keywords:
                if kw in col_clean:
                    mapped['num_col'] = col
                    break
        if not mapped['name_col']:
            for kw in name_keywords:
                if kw in col_clean:
                    mapped['name_col'] = col
                    break
        if not mapped['content_col']:
            for kw in content_keywords:
                if kw in col_clean:
                    mapped['content_col'] = col
                    break
        if not mapped['area_col']:
            for kw in area_keywords:
                if kw in col_clean:
                    mapped['area_col'] = col
                    break
        if not mapped['grade_col']:
            for kw in grade_keywords:
                if kw in col_clean:
                    mapped['grade_col'] = col
                    break

    if not mapped['area_col']:
        for col in columns:
            if col not in [mapped['num_col'], mapped['name_col'], mapped['content_col']]:
                sample_vals = df_processed[col].dropna().astype(str).tolist()[:50]
                if any(v in ['자율활동', '동아리활동', '진로활동', '자율', '동아리', '진로'] for v in sample_vals):
                    mapped['area_col'] = col
                    break

    if not mapped['num_col'] and len(columns) > 0:
        mapped['num_col'] = columns[0]
    if not mapped['name_col'] and len(columns) > 1:
        mapped['name_col'] = columns[1]
    if not mapped['content_col'] and len(columns) > 2:
        sample = df_processed.head(10)
        max_len_col = columns[-1]
        max_len = 0
        for c in columns:
            if c not in [mapped['num_col'], mapped['name_col']]:
                avg_len = sample[c].astype(str).str.len().mean()
                if avg_len > max_len:
                    max_len = avg_len
                    max_len_col = c
        mapped['content_col'] = max_len_col

    mapped['extra_cols'] = [
        c for c in columns 
        if c not in [mapped['num_col'], mapped['name_col'], mapped['content_col']]
        and not str(c).strip().lower().startswith('nan')
        and not str(c).strip().lower().startswith('unnamed')
        and str(c).strip() != 'None'
    ]
    return df_processed, mapped


def extract_current_grade_class(df_raw: pd.DataFrame) -> tuple:
    """
    엑셀 상단 셀(B8 기본, B6~B8 범위 및 상단 셀)에서 현재 학년/반 정보를 추출합니다.
    Returns: (grade: int, ban: int)
    """
    if df_raw is None or df_raw.empty:
        return 3, 1

    check_coords = [(7, 1), (5, 1), (6, 1), (4, 1), (3, 1), (8, 1)]
    for r, c in check_coords:
        if r < len(df_raw) and c < df_raw.shape[1]:
            val = safe_str(df_raw.iloc[r, c])
            m = re.search(r'(\d+)\s*학년\s*(\d+)\s*반', val)
            if m:
                return int(m.group(1)), int(m.group(2))

    for r in range(min(10, len(df_raw))):
        for c in range(min(6, df_raw.shape[1])):
            val = safe_str(df_raw.iloc[r, c])
            m = re.search(r'(\d+)\s*학년\s*(\d+)\s*반', val)
            if m:
                return int(m.group(1)), int(m.group(2))

    return 3, 1


def detect_actual_record_type(filename: str, df_raw: pd.DataFrame) -> str:
    """
    NEIS XLS 표준 셀 위치 (창체: D4, 세특: H6, 행특: E4) 및 내부 헤더/셀 값을 분석하여
    데이터 유형("창체", "세특", "행특")을 100% 정밀 판별합니다.
    """
    def get_val(r, c):
        if r < len(df_raw) and c < df_raw.shape[1]:
            v = df_raw.iloc[r, c]
            if safe_notna(v):
                return safe_str(v).replace(" ", "")
        return ""

    # D4 (r=3, c=3) -> 창체
    d4_val = get_val(3, 3)
    if '창의적체험활동상황' in d4_val or '창의적체험활동' in d4_val:
        return "창체"

    # H6 (r=5, c=7) -> 세특
    h6_val = get_val(5, 7)
    if '세부능력' in h6_val or '세부능력및특기사항' in h6_val:
        return "세특"

    # E4 (r=3, c=4) -> 행특
    e4_val = get_val(3, 4)
    if '행동특성' in e4_val or '행동특성및종합의견' in e4_val:
        return "행특"

    # NEIS 기존 표준 위치 보완
    cell_g3 = get_val(2, 6)
    cell_d2 = get_val(1, 3)
    cell_c2 = get_val(1, 2)
    if '세부능력' in cell_g3 or '학교생활기록부세부능력' in cell_g3:
        return "세특"
    if '행동특성' in cell_d2 or '학교생활기록부행동특성' in cell_d2:
        return "행특"
    if '창의적체험활동' in cell_c2 or '학교생활기록부창의적체험활동' in cell_c2:
        return "창체"

    # 상단 10개 행 전수 조사 (셀 위치 오차나 병합 셀 보완)
    top_rows_text = ""
    for r in range(min(10, len(df_raw))):
        row_str = "".join([safe_str(v).replace(" ", "") for v in df_raw.iloc[r].values if safe_notna(v)])
        top_rows_text += row_str + " "

    if '학교생활기록부창의적체험활동' in top_rows_text or '창의적체험활동상황' in top_rows_text or '영역별특기사항' in top_rows_text:
        return "창체"
    if '학교생활기록부행동특성' in top_rows_text or '행동특성및종합의견' in top_rows_text:
        return "행특"
    if '학교생활기록부세부능력' in top_rows_text or '세부능력및특기사항' in top_rows_text:
        return "세특"

    df_processed, col_map = detect_columns(df_raw)
    all_col_names = [safe_str(c).replace(" ", "") for c in df_processed.columns] + [safe_str(c).replace(" ", "") for c in df_raw.columns]
    col_str_joined = "".join(all_col_names)

    area_vals = []
    if col_map.get('area_col') and col_map['area_col'] in df_processed.columns:
        area_vals = [safe_str(v).strip() for v in df_processed[col_map['area_col']].dropna().tolist()[:50]]
    else:
        for col in df_processed.columns:
            vals = [safe_str(v).strip() for v in df_processed[col].dropna().tolist()[:30]]
            if any(v in ['자율활동', '동아리활동', '진로활동', '자율', '동아리', '진로', '봉사활동', '개별영역'] for v in vals):
                area_vals = vals
                break

    is_changche_area = any(v in ['자율활동', '동아리활동', '진로활동', '자율', '동아리', '진로', '봉사활동', '개별영역'] for v in area_vals)
    if '창의적체험활동' in top_rows_text or '창체' in top_rows_text or is_changche_area:
        return "창체"

    if '행동특성' in top_rows_text or '종합의견' in top_rows_text or '행동특성' in col_str_joined or '종합의견' in col_str_joined:
        return "행특"

    if '세부능력' in top_rows_text or '과목명' in top_rows_text or '교과' in top_rows_text or '세특' in top_rows_text or '세부능력' in col_str_joined or '과목' in col_str_joined:
        return "세특"

    return "세특"


def is_header_or_footer_row(row_dict: dict, num_col: str, name_col: str) -> bool:
    """
    NEIS 페이지 꼬리말(페이지번호/학교명), 페이지 상단 제목(학년 반), 표 헤더 행을 감지합니다.
    """
    vals = [safe_str(v) for v in row_dict.values() if safe_notna(v)]
    vals_str = ' '.join(vals)
    if not vals_str.strip():
        return True
    # 꼬리말 (예: 5.0 / 309.0 포곡고등학교)
    if re.search(r'\d+\s*/\s*\d+', vals_str) or '포곡고등학교' in vals_str or '고등학교' in vals_str:
        return True
    # 페이지 상단 학년반 제목 (예: 3학년 2반)
    if re.search(r'^\d+\s*학년\s*\d+\s*반', vals_str.strip()):
        return True
    # 표 헤더 (예: 번 호, 성  명, 학 년)
    if '번 호' in vals_str or '성  명' in vals_str or '성 명' in vals_str or '세부능력 및 특기사항' in vals_str:
        num_v = safe_str(row_dict.get(num_col, ''))
        name_v = safe_str(row_dict.get(name_col, ''))
        if num_v in ['번 호', '번호'] or name_v in ['성  명', '성명', '이름']:
            return True
    return False


# ==============================================================================
# 5. 지능형 페이지 파싱 엔진 (Core Refinement Engine)
# ==============================================================================
def refine_student_records(df: pd.DataFrame, col_map: dict) -> tuple:
    """
    페이지 나눔(NEIS 인쇄 경계)으로 인해 쪼개진 행들을 파싱합니다.
    가장 가까운 윗행 중 번호와 이름이 명시된 행(Active Record)을 추적하여,
    분할된 텍스트를 해당 학생 레코드의 문장 끝에 정확하게 이어 붙입니다.
    """
    num_col = col_map['num_col']
    name_col = col_map['name_col']
    content_col = col_map['content_col']
    area_col = col_map.get('area_col')
    grade_col = col_map.get('grade_col')

    refined_rows = []
    merge_logs = []
    hope_blank_records = []
    current_student_record = None
    active_num = ""
    active_name = ""
    active_hours = ""
    active_area = ""
    active_grade = ""
    active_cycle = "일반 영역"

    for idx, row in df.iterrows():
        try:
            row_dict = row.to_dict()

            # 섹션 헤더 (예: <진로 선택 과목>, <체육ㆍ예술>) 감지 시 이전 학생 레코드를 플러시하고 새로운 사이클로 전환
            row_vals_str = ' '.join([safe_str(v) for v in row_dict.values() if safe_notna(v)])
            sec_match = re.search(r'<([^>]+)>', row_vals_str)
            if sec_match:
                if current_student_record is not None:
                    if safe_str(current_student_record.get(content_col)):
                        refined_rows.append(current_student_record)
                    current_student_record = None
                raw_cyc = sec_match.group(1).strip()
                active_cycle = f"<{raw_cyc}>" if '<' not in raw_cyc else raw_cyc
                active_num = ""
                active_name = ""
                active_grade = ""
                continue

            if is_header_or_footer_row(row_dict, num_col, name_col):
                continue

            num_val = safe_val(row[num_col]) if num_col and num_col in row else None
            name_val = safe_val(row[name_col]) if name_col and name_col in row else None
            content_val = clean_text_content(safe_val(row[content_col])) if content_col and content_col in row else ""
            if not content_val or len(content_val) < 10:
                alt_texts = []
                for c in row.index:
                    if c not in [num_col, name_col, grade_col, '시간', '시 간', '이수시간']:
                        v_str = clean_text_content(safe_val(row[c]))
                        if len(v_str) > 15 and v_str.lower() not in ['nan', 'none']:
                            alt_texts.append(v_str)
                if alt_texts:
                    content_val = '\n\n'.join(alt_texts)

            area_val = safe_str(row[area_col]) if area_col and area_col in row else ""
            grade_val = safe_str(row[grade_col]).replace(".0", "") if grade_col and grade_col in row else ""

            hours_val = ""
            for h_col in ['시간', '시 간', '이수시간']:
                if h_col in row:
                    h_v = safe_val(row[h_col])
                    if safe_notna(h_v) and safe_str(h_v) not in ['', 'nan', 'NaN', 'None']:
                        hours_val = safe_str(h_v).replace(".0", "")
                        break

            is_num_empty = safe_isna(num_val) or safe_str(num_val) in ['', 'nan', 'NaN', 'None']
            is_name_empty = safe_isna(name_val) or safe_str(name_val) in ['', 'nan', 'NaN', 'None']

            num_str = "" if is_num_empty else safe_str(num_val)
            name_str = "" if is_name_empty else safe_str(name_val)

            if grade_val and grade_val.lower() != 'none':
                active_grade = grade_val
            if hours_val:
                active_hours = hours_val
            if area_val and area_val.lower() != 'none' and area_val not in ['영역', '활동영역', '창의적체험활동', '구분', '세부']:
                active_area = area_val

            if num_str and name_str:
                if active_num and num_str != active_num:
                    active_grade = "1"  # 새로운 학생이 시작되면 기본 1학년으로 초기화
                active_num = num_str
                active_name = name_str

            target_num = num_str if num_str else active_num
            target_name = name_str if name_str else active_name

            if not target_num or not target_name or not content_val:
                continue

            eff_grade = grade_val if (grade_val and grade_val.lower() != 'none') else (active_grade if active_grade else "1")
            eff_grade_clean = eff_grade.replace("학년", "").replace(".0", "").strip()

            # 이전 학생의 레코드가 남아있는데 대상 학생이 바뀐 경우 플러시!
            if current_student_record is not None:
                curr_n = safe_str(current_student_record.get(num_col))
                curr_nm = safe_str(current_student_record.get(name_col))
                curr_g = safe_str(current_student_record.get(grade_col, '')).replace("학년", "").replace(".0", "").strip() if grade_col else ""
                curr_cyc = current_student_record.get('cycle', '')
                
                # 학생이 바뀌었거나 사이클이 바뀌었거나, 엑셀에 명시된 학년이 변경된 경우 분리
                if (curr_n != target_num or curr_nm != target_name) or (curr_cyc != active_cycle) or (grade_val and grade_val.replace("학년", "").strip() != curr_g):
                    if safe_str(current_student_record.get(content_col)):
                        refined_rows.append(current_student_record)
                    current_student_record = None

            # NEIS 진로활동 '희망사항' (Column I/F) 및 학생 희망진로 내용 (Column M) 검사
            is_hope_row, hope_val = check_jinro_hope_field(row_dict, row)
            if is_hope_row:
                active_area = "진로활동"
                # Column F가 '희망분야'이고 Column G(학생 희망분야)가 비어있는 경우에만 정밀 오류 기록
                if not hope_val and target_num and target_name:
                    hope_blank_records.append({
                        "student_id": target_num,
                        "student_name": target_name,
                        "category": "창체",
                        "taken_grade": f"{eff_grade_clean}학년" if eff_grade_clean else "3학년",
                        "sub_category": "진로활동",
                        "original_text": "희망분야 (공란)",
                        "suggested_text": "학생 희망분야 기입",
                        "reason": "진로활동 희망분야가 공란으로 기재되어 있으니 학생의 희망분야를 확인하여 기입하십시오.",
                        "severity": "수정 필수"
                    })
                # 희망분야 헤더 행은 서술문이 아니므로 원문 텍스트 추출에서는 제외
                continue

            if not content_val:
                continue

            curr_num = current_student_record[num_col] if current_student_record else None
            curr_name = current_student_record[name_col] if current_student_record else None
            curr_area = current_student_record.get('sub_category', current_student_record.get(area_col, '')) if current_student_record else ''
            curr_grade = safe_str(current_student_record.get(grade_col, '')).replace("학년", "").replace(".0", "").strip() if (current_student_record and grade_col) else ""

            is_same_student = (current_student_record is not None and curr_num == target_num and curr_name == target_name)
            is_same_grade = (not grade_val or grade_val.replace("학년", "").strip() == curr_grade)

            if area_val and area_val.lower() != 'none' and area_val not in ['영역', '활동영역', '창의적체험활동', '구분', '세부']:
                effective_area = area_val
            elif active_area:
                effective_area = active_area
            elif is_same_student and curr_area:
                effective_area = curr_area
            elif area_col:
                effective_area = get_changche_area_from_cell(area_val, content_val)
            else:
                effective_area = ""

            is_same_area = (not area_col or not effective_area or not curr_area or effective_area == curr_area or effective_area in curr_area or curr_area in effective_area)

            # 스포츠클럽은 별도의 독립 서술문 영역이 아니라 해당 학년 동아리의 부속 이수 기록임
            is_sports_club = '스포츠클럽' in content_val
            is_new_activity_start = (not is_sports_club) and bool(re.match(
                r'^(?:(?:\([12]학기\))?\s*[가-힣a-zA-Z0-9\s·/Ⅰ-Ⅻ()\-_]{1,20}\s*[:：]|\([가-힣a-zA-Z0-9\s·/]+\)\s*\(\d+시간\))',
                content_val
            ))

            # 같은 학생 + 같은 학년 + 같은 영역일 때만 이전 레코드에 이어붙이기 수행
            if is_same_student and is_same_grade and is_same_area and not is_new_activity_start:
                prev_content = current_student_record[content_col]
                merged_content = smart_concatenate_text(prev_content, content_val)
                current_student_record[content_col] = merged_content
                current_student_record['_merged_count'] += 1
                current_student_record['_merged_rows'].append(idx + 2)

                merge_logs.append({
                    'excel_row': idx + 2,
                    'target_student': f"{current_student_record[num_col]}번 {current_student_record[name_col]}",
                    'appended_text': content_val,
                    'result_content_snippet': merged_content[-80:]
                })
                continue

            if current_student_record is not None:
                if str(current_student_record.get(content_col, '')).strip():
                    refined_rows.append(current_student_record)

            new_record = row.to_dict()
            new_record[num_col] = target_num
            new_record[name_col] = target_name
            new_record[content_col] = content_val
            new_record['cycle'] = active_cycle

            eff_area = effective_area

            new_record['sub_category'] = eff_area
            if area_col:
                new_record[area_col] = eff_area
            else:
                new_record['영역'] = eff_area
            new_record['sub_category'] = eff_area
            if grade_col:
                new_record[grade_col] = f"{eff_grade_clean}학년" if eff_grade_clean else "3학년"

            # 시간(이수시간) 컬럼 보전 및 계승
            for h_col in ['시간', '시 간', '이수시간']:
                if h_col in df.columns or h_col in new_record:
                    if pd.isna(new_record.get(h_col)) or str(new_record.get(h_col, '')).strip() in ['', 'nan', 'NaN', 'None']:
                        new_record[h_col] = active_hours

            new_record['_original_excel_row'] = idx + 2
            new_record['_merged_count'] = 0
            new_record['_merged_rows'] = []

            current_student_record = new_record

        except Exception as e:
            st.error(f"행 {idx + 2}번 처리 중 예외 발생: {str(e)}")
            continue

    if current_student_record is not None and str(current_student_record.get(content_col, '')).strip():
        refined_rows.append(current_student_record)

    # Post-Processing Pass: Merge cut-off fragment rows for same student + same grade
    final_refined_rows = []
    for r in refined_rows:
        if not final_refined_rows:
            final_refined_rows.append(r)
            continue
        
        last_r = final_refined_rows[-1]
        same_st = (str(last_r[num_col]).strip() == str(r[num_col]).strip() and str(last_r[name_col]).strip() == str(r[name_col]).strip())
        same_gd = (str(last_r.get(grade_col, '')).replace('학년', '').strip() == str(r.get(grade_col, '')).replace('학년', '').strip())
        last_sub = str(last_r.get('sub_category', last_r.get(area_col, '') if area_col else '')).strip()
        curr_sub = str(r.get('sub_category', r.get(area_col, '') if area_col else '')).strip()
        same_ar = (bool(last_sub and curr_sub) and (last_sub == curr_sub or curr_sub in last_sub or last_sub in curr_sub))
        
        last_text = str(last_r[content_col]).strip()
        curr_text = str(r[content_col]).strip()
        
        is_new_header = bool(re.match(
            r'^(?:(?:\([12]학기\))?\s*[가-힣a-zA-Z0-9\s·/Ⅰ-Ⅻ()\-_]{1,20}\s*[:：]|\([가-힣a-zA-Z0-9\s·/]+\)\s*\(\d+시간\))',
            curr_text
        ))

        # 반드시 같은 학생 + 같은 학년일 때만 끊긴 문장 이어붙이기 허용!
        if same_st and same_gd and same_ar and not is_new_header:
            merged = smart_concatenate_text(last_text, curr_text)
            last_r[content_col] = merged
            last_r['_merged_count'] += 1 + r.get('_merged_count', 0)
            if '_merged_rows' in r:
                last_r['_merged_rows'].extend(r['_merged_rows'])
            merge_logs.append({
                'excel_row': r.get('_original_excel_row', 0),
                'target_student': f"{last_r[num_col]}번 {last_r[name_col]}",
                'appended_text': curr_text,
                'result_content_snippet': merged[-80:]
            })
        else:
            final_refined_rows.append(r)

    refined_df = pd.DataFrame(final_refined_rows)
    return refined_df, merge_logs, hope_blank_records


# ==============================================================================
# 6. 세특 과목명 및 학기 정보 자동 분리 로직 (Regex Engine)
# ==============================================================================
def clean_subject_name(raw_subj: str) -> str:
    """
    과목명에서 '(1학기)', '(2학기)', '(1 학기)', '(2 학기)' 등 학기 접두어를 제거하여
    순수 과목명만 정제 추출합니다. (예: '(1학기) 국어' -> '국어')
    """
    if not raw_subj:
        return ""
    cleaned = safe_str(raw_subj).strip()
    cleaned = re.sub(r'^\([12]\s*학기\)\s*', '', cleaned).strip()
    return cleaned


def split_subject_details(df: pd.DataFrame, col_map: dict) -> pd.DataFrame:
    if df.empty:
        return df

    num_col = col_map['num_col']
    name_col = col_map['name_col']
    content_col = col_map['content_col']
    extra_cols = col_map['extra_cols']

    # Improved regex to find subject headings (including multi-word subjects like '화법과 작문', '윤리와 사상')
    # Match starts at string start, newline, after sentence ending punctuation (. ! ?), or after closing paren )
    pattern = r'(?:^|\n|[.!?]\s*|\)\s*)\s*(((?:\([12]\s*학기\))?\s*[가-힣a-zA-Z0-9\s·/Ⅰ-Ⅻ()\-_]{1,20}))\s*[:：]'
    
    ignore_hdr_keywords = [
        '주제', '탐구주제', '탐구 주제', '결과', '활동결과', '활동 결과', 
        '느낀점', '느낀 점', '소감', '의견', '참고문헌', '참고 문헌', '내용', '활동내용',
        '일시', '장소', '대상', '목표', '방법', '동기', '탐구 동기', '탐구동기', '역할', '분야'
    ]

    unfolded_rows = []

    for _, row in df.iterrows():
        try:
            raw_text = safe_str(row.get(content_col, ''))
            if not raw_text:
                continue

            # 1. 줄바꿈 기호 표준화 (Alt+Enter, HTML <br>, \r\n -> \n)
            text_norm = raw_text.replace('<br/>', '\n').replace('<br>', '\n').replace('\r\n', '\n').replace('\r', '\n')
            
            base_info = {
                num_col: safe_val(row.get(num_col)),
                name_col: safe_val(row.get(name_col)),
                'cycle': row.get('cycle', '일반 영역')
            }
            for c in extra_cols:
                if c not in ['_merged_count', '_merged_rows', '_original_excel_row']:
                    base_info[c] = safe_val(row.get(c))

            row_subj = safe_str(row.get('과목명', row.get('과목', '')))
            clean_row_subj = clean_subject_name(row_subj)
            if not clean_row_subj or clean_row_subj.lower() in ['nan', 'none', ''] or clean_row_subj in ['세부능력및특기사항', '세부능력 및 특기사항', '세특', '세부']:
                clean_row_subj = '개세특'

            matches = []
            for m in re.finditer(pattern, text_norm):
                full_hdr = m.group(1).strip()
                
                # Clean heading & extract semester tag
                sem_m = re.search(r'\(([12]\s*학기)\)', full_hdr)
                if sem_m:
                    sem_tag = f"({sem_m.group(1).replace(' ', '')})"
                    clean_cand = full_hdr[sem_m.end():].strip()
                    clean_cand = re.sub(r'^[^\w가-힣]+', '', clean_cand).strip()
                else:
                    sem_tag = ""
                    clean_cand = clean_subject_name(full_hdr)

                byte_len = len(clean_cand.encode('euc-kr', errors='ignore'))
                
                # Exclude common body colon phrases
                if clean_cand in ignore_hdr_keywords or re.match(r'^\d+[\.\)]\s*', clean_cand):
                    continue

                if byte_len <= 30 and len(clean_cand) >= 1:
                    matches.append({
                        'start_full': m.start(1),
                        'start_hdr': m.start(1),
                        'end_hdr': m.end(),
                        'hdr_text': full_hdr,
                        'sem_tag': sem_tag,
                        'clean_subj': clean_cand
                    })

            grade_col_name = col_map.get('grade_col', '학년')

            if not matches:
                # 같은 학생 + 같은 학년의 이전 레코드가 있고, 이전 문장이 완전 종결되지 않고 중간에 잘린 경우만 이어붙이기 수행
                is_same_st_g = (unfolded_rows and 
                                unfolded_rows[-1].get(num_col) == base_info.get(num_col) and 
                                unfolded_rows[-1].get(name_col) == base_info.get(name_col) and
                                unfolded_rows[-1].get(grade_col_name) == base_info.get(grade_col_name))
                
                last_text = unfolded_rows[-1]['내용'].strip() if is_same_st_g else ''
                last_ended_sentence = bool(re.search(r'[.!?]|함$|임$|됨$|음$|함\.$|임\.$|됨\.$|음\.$', last_text))

                if is_same_st_g and not last_ended_sentence:
                    last_unfolded = unfolded_rows[-1]
                    merged_content = smart_concatenate_text(last_unfolded['내용'], text_norm)
                    last_unfolded['내용'] = merged_content
                    last_unfolded['글자수'] = len(merged_content)
                else:
                    item = base_info.copy()
                    item['과목명'] = clean_row_subj if clean_row_subj not in ['세부능력및특기사항', '세부능력 및 특기사항', '세특', '세부', 'nan'] else '개세특'
                    item['내용'] = text_norm.replace('\n', ' ').strip()
                    item['글자수'] = len(item['내용'])
                    unfolded_rows.append(item)
                continue

            # If there is content before the first matched subject heading
            first_m = matches[0]
            if first_m['start_full'] > 0:
                pre_content = text_norm[:first_m['start_full']].strip()
                if len(pre_content) > 0:
                    is_same_st_g = (unfolded_rows and 
                                    unfolded_rows[-1].get(num_col) == base_info.get(num_col) and 
                                    unfolded_rows[-1].get(name_col) == base_info.get(name_col) and
                                    unfolded_rows[-1].get(grade_col_name) == base_info.get(grade_col_name))
                    last_text = unfolded_rows[-1]['내용'].strip() if is_same_st_g else ''
                    last_ended_sentence = bool(re.search(r'[.!?]|함$|임$|됨$|음$|함\.$|임\.$|됨\.$|음\.$', last_text))

                    if is_same_st_g and not last_ended_sentence:
                        last_unfolded = unfolded_rows[-1]
                        merged_content = smart_concatenate_text(last_unfolded['내용'], pre_content)
                        last_unfolded['내용'] = merged_content
                        last_unfolded['글자수'] = len(merged_content)
                    else:
                        item = base_info.copy()
                        item['과목명'] = clean_row_subj if clean_row_subj not in ['세부능력및특기사항', '세부능력 및 특기사항', '세특', '세부', 'nan'] else '개세특'
                        item['내용'] = pre_content.replace('\n', ' ').strip()
                        item['글자수'] = len(item['내용'])
                        unfolded_rows.append(item)

            curr_cycle = base_info.get('cycle', '일반 영역')

            for i in range(len(matches)):
                m_curr = matches[i]
                start_content = m_curr['end_hdr']
                end_content = matches[i+1]['start_full'] if i + 1 < len(matches) else len(text_norm)
                
                snippet = text_norm[start_content:end_content].strip()
                sem_tag = m_curr['sem_tag']
                clean_s = m_curr['clean_subj'] if m_curr['clean_subj'] else clean_row_subj
                if clean_s in ['세부능력및특기사항', '세부능력 및 특기사항', '세특', '세부', 'nan']:
                    clean_s = '개세특'
                
                if sem_tag:
                    full_subj_name = f"{sem_tag} {clean_s}"
                else:
                    full_subj_name = clean_s
                
                # 개세특은 일반 영역에서 학년의 가장 마지막 과목 마침표 뒤 \n\n 이후 문단에서만 최대 1개 추출
                is_last_subj_match = (i == len(matches) - 1)
                
                if is_last_subj_match and curr_cycle == '일반 영역':
                    parts = re.split(r'\n\s*\n', snippet)
                    main_text = parts[0].replace('\n', ' ').strip()
                    extra_text = '\n\n'.join(parts[1:]).replace('\n', ' ').strip() if len(parts) > 1 else ""

                    if main_text:
                        item = base_info.copy()
                        item['과목명'] = full_subj_name
                        item['내용'] = main_text
                        item['글자수'] = len(main_text)
                        unfolded_rows.append(item)

                    if extra_text and len(extra_text) > 10 and not re.match(r'^[가-힣a-zA-Z0-9\s·/Ⅰ-Ⅻ()\-_]{1,20}[:：]', extra_text):
                        item = base_info.copy()
                        item['과목명'] = '개세특'
                        item['내용'] = extra_text
                        item['글자수'] = len(extra_text)
                        if '이수학년' in base_info:
                            item['이수학년'] = base_info['이수학년']
                        unfolded_rows.append(item)
                else:
                    clean_text = snippet.replace('\n', ' ').strip()
                    if clean_text:
                        item = base_info.copy()
                        item['과목명'] = full_subj_name
                        item['내용'] = clean_text
                        item['글자수'] = len(clean_text)
                        unfolded_rows.append(item)

        except Exception as e:
            st.error(f"세특 과목 분리 처리 중 예외 발생 ({row.get(name_col, '')}): {str(e)}")

    res_df = pd.DataFrame(unfolded_rows)
    if res_df.empty:
        return res_df

    # 지침 규격(xls 파일 형식.md): 개세특은 학생별 학년당 최대 1개만 허용 (0개 또는 1개)
    # 중간 과목에서 발생한 엔터 2번 문단은 해당 과목의 계속되는 내용으로 병합하고, 학년의 맨 마지막 문단 개세특만 1개 보존
    grade_col_name = col_map.get('grade_col', '학년')
    res_df['norm_grade'] = res_df[grade_col_name].astype(str).str.replace('학년', '').str.replace('.0', '').str.strip() + '학년'
    cleaned_rows = []
    
    grouped = res_df.groupby([num_col, name_col], sort=False)
    for (st_num, st_name), st_group in grouped:
        for cyc_name, cyc_group in st_group.groupby('cycle', sort=False):
            if cyc_name != '일반 영역':
                cyc_recs = []
                for r in cyc_group.to_dict('records'):
                    if r.get('과목명') == '개세특':
                        if cyc_recs:
                            prev_r = cyc_recs[-1]
                            prev_r['내용'] = prev_r['내용'].strip() + ' ' + r['내용'].strip()
                            prev_r['글자수'] = len(prev_r['내용'])
                    else:
                        cyc_recs.append(r)
                cleaned_rows.extend(cyc_recs)
                continue
            
            for g_val, g_group in cyc_group.groupby('norm_grade', sort=False):
                records = g_group.to_dict('records')
                new_g_recs = []
                for i, r in enumerate(records):
                    if r.get('과목명') == '개세특':
                        is_very_last = (i == len(records) - 1)
                        if is_very_last:
                            new_g_recs.append(r)
                        else:
                            if new_g_recs:
                                prev_r = new_g_recs[-1]
                                prev_r['내용'] = prev_r['내용'].strip() + ' ' + r['내용'].strip()
                                prev_r['글자수'] = len(prev_r['내용'])
                    else:
                        new_g_recs.append(r)
                cleaned_rows.extend(new_g_recs)

    final_res = pd.DataFrame(cleaned_rows)
    if 'norm_grade' in final_res.columns:
        final_res = final_res.drop(columns=['norm_grade'])
    return final_res


SUBJECT_GRADE_MAP = {
    '국어': '1학년', '(1학기) 수학': '1학년', '(2학기) 수학': '1학년', '수학': '1학년', '영어': '1학년',
    '한국사': '1학년', '통합사회': '1학년', '(1학기) 통합과학': '1학년', '(2학기) 통합과학': '1학년', '통합과학': '1학년',
    '과학탐구실험': '1학년', '기술·가정': '1학년', '기술가정': '1학년', '(1학기) 체육': '1학년', '(2학기) 체육': '1학년',
    '체육': '1학년', '음악': '1학년', '미술': '1학년',
    '독서': '2학년', '문학': '2학년', '수학Ⅰ': '2학년', '수학Ⅱ': '2학년', '영어Ⅰ': '2학년', '영어Ⅱ': '2학년',
    '물리학Ⅰ': '2학년', '화학Ⅰ': '2학년', '생명과학Ⅰ': '2학년', '지구과학Ⅰ': '2학년', '정보': '2학년',
    '일본어Ⅰ': '2학년', '중국어Ⅰ': '2학년', '한문Ⅰ': '2학년', '운동과 건강': '2학년', '음악 연주': '2학년', '기하': '2학년',
    '세계지리': '2학년', '세계사': '2학년', '동아시아사': '2학년', '정치와 법': '2학년', '윤리와 사상': '2학년', '경제': '2학년',
    '화법과 작문': '3학년', '미적분': '3학년', '확률과 통계': '3학년', '영어 독해와 작문': '3학년', '언어와 매체': '3학년',
    '교육학': '3학년', '사회문제 탐구': '3학년', '화학Ⅱ': '3학년', '생명과학Ⅱ': '3학년', '지구과학Ⅱ': '3학년',
    '물리학Ⅱ': '3학년', '생활과 과학': '3학년', '스포츠 생활': '3학년', '공학 일반': '3학년', '인공지능 기초': '3학년',
    '사회·문화': '3학년', '생활과 윤리': '3학년', '실용 경제': '3학년', '진로 영어': '3학년', '영어 회화': '3학년',
    '일본어 회화Ⅰ': '3학년', '미술 창작': '2학년'
}


def extract_row_grade(row: pd.Series, c_map: dict, num_str: str, sub_cat: str = "") -> int:
    """
    행 레코드에서 이수학년(1학년, 2학년, 3학년)을 추출합니다.
    표준 과목 맵 및 엑셀 파일 내부의 '학년' 열 값 및 계승된 행 서식 데이터를 100% 우선 적용합니다.
    """
    row_dict = row.to_dict()

    # 0. 표준 과목 학년 맵 탐색 (최우선)
    subj_clean = clean_subject_name(sub_cat)
    if sub_cat in SUBJECT_GRADE_MAP:
        return int(SUBJECT_GRADE_MAP[sub_cat].replace('학년', ''))
    if subj_clean in SUBJECT_GRADE_MAP:
        return int(SUBJECT_GRADE_MAP[subj_clean].replace('학년', ''))

    # 1. 엑셀 행/레코드의 학년 열 지정 값 검사
    grade_col_candidates = [
        c for c in row.index 
        if any(k in str(c).replace(" ", "").lower() for k in ['학년', 'grade', '이수학년'])
        and not str(c).startswith('_')
    ]
    
    for gc in grade_col_candidates:
        val = str(row_dict.get(gc, '')).strip()
        m = re.search(r'([1-3])\s*학년', val)
        if m:
            return int(m.group(1))
        m_num = re.search(r'^[1-3]$', val)
        if m_num:
            return int(m_num.group())

    # 2. 행 텍스트 내 학년 표기 검사
    row_text = " ".join([safe_str(v) for k, v in row_dict.items() if safe_notna(v) and not str(k).startswith('_')])
    full_text = f"{sub_cat} {row_text}"

    m_text = re.search(r'([1-3])\s*학년', full_text)
    if m_text:
        return int(m_text.group(1))

    # 3. 5자리 학번(예: 30101 -> 3학년)의 첫 번째 자릿수 검사
    if len(num_str) == 5 and num_str[0] in ['1', '2', '3']:
        return int(num_str[0])

    return 3


def auto_detect_current_grade(data_store: dict) -> int:
    """
    파일명(무작위 알파벳 등)에 전혀 의존하지 않고, 엑셀 파일 내부의 학번 데이터,
    헤더 셀 및 본문 교과목 텍스트만을 정밀 정독 분석하여 학생들의 대상 현재 학년을 100% 자동 감지합니다.
    """
    max_detected = 1

    for t_key in ["창체", "세특", "행특"]:
        if t_key not in data_store or data_store[t_key] is None:
            continue
        df = data_store[t_key]['df']
        c_map = data_store[t_key]['col_map']
        num_c = c_map.get('num_col')

        if num_c and num_c in df.columns:
            num_vals = [re.sub(r'\D', '', str(v)) for v in df[num_c].values if pd.notna(v)]
            five_digit_ids = [v for v in num_vals if len(v) == 5]
            for s_id in five_digit_ids:
                first_digit = int(s_id[0])
                if first_digit in [1, 2, 3]:
                    max_detected = max(max_detected, first_digit)

        row_text_sample = " ".join([str(v) for v in df.astype(str).values.flatten()[:1000] if pd.notna(v)])
        
        grade3_keywords = [
            '3학년', '3 학년', '미적분', '기하', '화법과 작문', '화법과작문', '언어와 매체', '언어와매체',
            '물리학Ⅱ', '물리학2', '화학Ⅱ', '화학2', '생명과학Ⅱ', '생명과학2', '지구과학Ⅱ', '지구과학2',
            '영어 독해와 작문', '융합과학', '생활과 윤리', '윤리와 사상'
        ]
        if any(k in row_text_sample for k in grade3_keywords):
            return 3

        grade2_keywords = [
            '2학년', '2 학년', '문학', '독서', '수학Ⅰ', '수학1', '수학Ⅱ', '수학2',
            '영어Ⅰ', '영어1', '영어Ⅱ', '영어2', '물리학Ⅰ', '물리학1', '화학Ⅰ', '화학1',
            '생명과학Ⅰ', '생명과학1', '지구과학Ⅰ', '지구과학1', '한국지리', '세계지리', '동아시아사', '세계사', '경제', '정치와 법', '사회·문화'
        ]
        if any(k in row_text_sample for k in grade2_keywords):
            max_detected = max(max_detected, 2)

    return max(max_detected, 3)


def auto_detect_class_number(data_store: dict) -> int:
    """
    창체(B6), 세특(B8/B6), 행특(B6) 등 각 영역별 원본 raw_df 지정 셀 위치를 점검하여 반 번호를 자동 감지합니다.
    """
    for t_key in ["창체", "세특", "행특"]:
        if t_key not in data_store or data_store[t_key] is None:
            continue
        item = data_store[t_key]
        raw_df = item.get('raw_df')
        if raw_df is None and isinstance(data_store.get('raw_data'), dict):
            raw_df = data_store['raw_data'].get(t_key)
        if raw_df is None:
            raw_df = item.get('df')
        if raw_df is None or raw_df.empty:
            continue

        g_num, b_num = extract_current_grade_class(raw_df)
        if b_num:
            return b_num

    return 1


# ==============================================================================
# 7. 생기부 레코드 리스트 패킹 헬퍼
# ==============================================================================
def prepare_records_for_llm(data_store: dict, target_current_grade: int = None) -> list:
    """
    LLM API에 전달할 [학번, 이름, 현재학년, 구분, 이수학년, 세부, 이수시간, 기록 텍스트] 페이로드를 생성합니다.
    """
    if target_current_grade is None:
        target_current_grade = auto_detect_current_grade(data_store)

    ban_num = auto_detect_class_number(data_store)

    student_max_grades = {}
    raw_records = []

    for t_key in ["창체", "세특", "행특"]:
        if t_key not in data_store or data_store[t_key] is None:
            continue
        item = data_store[t_key]
        df = item['df']
        c_map = item['col_map']
        num_c, name_c, content_c = c_map['num_col'], c_map['name_col'], c_map['content_col']

        for _, row in df.iterrows():
            num_raw = safe_str(row.get(num_c, ''))
            name_val = safe_str(row.get(name_c, ''))
            num_str = re.sub(r'\D', '', num_raw)

            text_content = safe_str(row.get('내용', row.get(content_c, '')))
            if t_key == "창체":
                raw_sub = safe_str(row.get('sub_category', row.get('영역', row.get('활동영역', ''))))
                sub_cat = get_changche_area_from_cell(raw_sub, text_content)
            elif t_key == "세특":
                sub_cat = safe_str(row.get('과목명', row.get('과목', '과목미지정')))
                if not sub_cat or sub_cat == 'nan':
                    sub_cat = "세부능력및특기사항"
            else:
                sub_cat = "행동특성"

            rec_grade = extract_row_grade(row, c_map, num_str, sub_cat)

            student_key = (name_val, num_str[-2:] if len(num_str) >= 2 else num_str)
            if student_key not in student_max_grades or rec_grade > student_max_grades[student_key]:
                student_max_grades[student_key] = rec_grade

            hours_val = ""
            for h_col in ['시간', '시 간', '이수시간']:
                if h_col in row:
                    h_v = safe_val(row[h_col])
                    if safe_notna(h_v) and safe_str(h_v) not in ['', 'nan', 'NaN', 'None']:
                        hours_val = safe_str(h_v).replace(".0", "")
                        break

            category_map = {"창체": "창체", "세특": "세특", "행특": "행발"}
            if text_content and text_content not in ['희망분야', '희망 분야', '희망분야:']:
                raw_records.append({
                    "num_str": num_str,
                    "name_val": name_val,
                    "category": category_map.get(t_key, t_key),
                    "taken_grade": f"{rec_grade}학년",
                    "taken_grade_num": rec_grade,
                    "sub_cat": sub_cat,
                    "hours": hours_val,
                    "text_content": text_content,
                    "student_key": student_key
                })

    records_payload = []
    for r in raw_records:
        s_key = r["student_key"]
        max_g = max(target_current_grade, student_max_grades.get(s_key, r["taken_grade_num"]))
        num_str = r["num_str"]

        if len(num_str) == 5:
            ban_part = num_str[1:3]
            num_part = num_str[3:]
            current_student_id = f"{max_g}{ban_part}{num_part}"
        elif len(num_str) in [1, 2]:
            current_student_id = f"{max_g}{ban_num:02d}{int(num_str):02d}"
        else:
            current_student_id = f"{max_g}{ban_num:02d}{num_str.zfill(2)[-2:]}"

        records_payload.append({
            "학번": current_student_id,
            "이름": r["name_val"],
            "현재학년": f"{max_g}학년",
            "구분": r["category"],
            "이수학년": r["taken_grade"],
            "세부": r["sub_cat"],
            "이수시간": r["hours"],
            "기록텍스트": r["text_content"]
        })

    return records_payload


# ==============================================================================
# 8. 서식 스타일 적용 엑셀 내보내기 함수 (Openpyxl - A4 가로인쇄, 15mm 여백)
# ==============================================================================
def create_audit_report_excel_bytes(audit_df: pd.DataFrame) -> bytes:
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AI_오탈자_검증_리포트"

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left = 0.59
    ws.page_margins.right = 0.59
    ws.page_margins.top = 0.59
    ws.page_margins.bottom = 0.59

    ws.oddFooter.right.text = "&P/&N"
    ws.evenFooter.right.text = "&P/&N"
    ws.views.sheetView[0].showGridLines = True

    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
    body_font = Font(name="맑은 고딕", size=10)
    font_red = Font(name="맑은 고딕", size=10, bold=True, color="990000")
    font_amber = Font(name="맑은 고딕", size=10, bold=True, color="92400E")

    fill_even = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_required = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    fill_recommended = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    headers = ["학년", "반", "번호", "학번", "이름", "구분", "이수학년", "세부", "이수시간", "원문 (수정 전)", "수정 후 (제안)", "수정 이유/근거", "수정구분"]
    ws.append(headers)

    ws.row_dimensions[1].height = 28
    for c_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    for row_idx, r in enumerate(audit_df.to_dict('records'), start=2):
        s_id = str(r.get("학번", r.get("학번(숫자5자리)", "00000")))
        grade = s_id[0] if len(s_id) == 5 and s_id[0].isdigit() else ""
        ban = str(int(s_id[1:3])) if len(s_id) == 5 and s_id[1:3].isdigit() else ""
        num = str(int(s_id[3:])) if len(s_id) == 5 and s_id[3:].isdigit() else ""

        row_values = [
            grade,
            ban,
            num,
            s_id,
            r.get("이름", ""),
            r.get("구분", ""),
            r.get("이수학년", ""),
            r.get("세부", ""),
            r.get("이수시간", r.get("시간", "")),
            r.get("수정전", ""),
            r.get("수정 후", ""),
            r.get("수정해야하는 이유나 근거", ""),
            r.get("수정구분", "")
        ]
        ws.append(row_values)

        row_fill = fill_even if row_idx % 2 == 0 else fill_odd
        mod_type = str(r.get("수정구분", ""))

        for col_idx in range(1, len(row_values) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = body_font
            cell.fill = row_fill
            cell.border = thin_border
            
            if col_idx in [1, 2, 3, 4, 5, 6, 7, 8, 9, 13]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

            if col_idx == 13:
                if "필수" in mod_type:
                    cell.fill = fill_required
                    cell.font = font_red
                elif "권장" in mod_type:
                    cell.fill = fill_recommended
                    cell.font = font_amber

    col_widths = {
        'A': 6, 'B': 6, 'C': 6, 'D': 9, 'E': 9, 'F': 8, 'G': 10, 'H': 12, 'I': 9,
        'J': 36,
        'K': 36,
        'L': 36,
        'M': 11
    }

    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    wb.save(output)
    return output.getvalue()


# ==============================================================================
# 9. PDF 인쇄용 내보내기 함수 (ReportLab - A4 가로인쇄, 15mm 여백, 페이지 꼬리말)
# ==============================================================================
def register_korean_font():
    font_paths = [
        ("Malgun", "C:\\Windows\\Fonts\\malgun.ttf"),
        ("Gulim", "C:\\Windows\\Fonts\\gulim.ttc"),
        ("Batang", "C:\\Windows\\Fonts\\batang.ttc"),
        ("NanumGothic", "/usr/share/fonts/truetype/nanum/NanumGothic.ttf"),
    ]
    for font_name, font_path in font_paths:
        if os.path.exists(font_path):
            try:
                if font_name not in pdfmetrics.getRegisteredFontNames():
                    pdfmetrics.registerFont(TTFont(font_name, font_path))
                return font_name
            except Exception:
                pass

    try:
        if 'HYGothic-Medium' not in pdfmetrics.getRegisteredFontNames():
            pdfmetrics.registerFont(UnicodeCIDFont('HYGothic-Medium'))
        return 'HYGothic-Medium'
    except Exception:
        pass

    try:
        if 'HYSMyeongJo-Medium' not in pdfmetrics.getRegisteredFontNames():
            pdfmetrics.registerFont(UnicodeCIDFont('HYSMyeongJo-Medium'))
        return 'HYSMyeongJo-Medium'
    except Exception:
        pass

    return "Helvetica"


class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(num_pages)
            super().showPage()
        super().save()

    def draw_page_number(self, page_count):
        self.saveState()
        font_name = register_korean_font()
        self.setFont(font_name, 9)
        self.setFillColor(colors.HexColor("#4B5563"))
        page_text = f"({self._pageNumber}/{page_count})"
        self.drawRightString(841.89 - 42.52, 20, page_text)
        self.restoreState()


def create_audit_report_pdf_bytes(audit_df: pd.DataFrame) -> bytes:
    font_name = register_korean_font()
    buffer = io.BytesIO()
    margin_pt = 42.52

    # Dynamic Title Construction (학교생활기록부 검증 리포트 - [구분] - [학년]학년 [반]반)
    cat_str = "통합"
    if '구분' in audit_df.columns and not audit_df['구분'].dropna().empty:
        unique_cats = [str(x).strip() for x in audit_df['구분'].dropna().unique() if str(x).strip()]
        if unique_cats:
            cat_str = " · ".join(unique_cats)

    grade_ban_str = ""
    id_col = '학번(숫자5자리)' if '학번(숫자5자리)' in audit_df.columns else ('학번' if '학번' in audit_df.columns else None)
    if id_col and not audit_df[id_col].dropna().empty:
        for val in audit_df[id_col].dropna().astype(str):
            s_id = val.strip()
            if len(s_id) == 5 and s_id.isdigit():
                g_num = s_id[0]
                b_num = int(s_id[1:3])
                grade_ban_str = f" - {g_num}학년 {b_num}반"
                break

    doc_title_text = f"학교생활기록부 검증 리포트 - {cat_str}{grade_ban_str}"
    
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=margin_pt,
        rightMargin=margin_pt,
        topMargin=margin_pt,
        bottomMargin=margin_pt + 15
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'PDFDocTitle',
        parent=styles['Heading1'],
        fontName=font_name,
        fontSize=13,
        leading=16,
        textColor=colors.HexColor("#1E3A8A"),
        spaceAfter=8
    )
    cell_style = ParagraphStyle(
        'PDFCellText',
        fontName=font_name,
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#1F2937")
    )
    header_cell_style = ParagraphStyle(
        'PDFHeaderCellText',
        fontName=font_name,
        fontSize=8,
        leading=10,
        textColor=colors.white,
        alignment=1
    )

    elements = []
    elements.append(Paragraph(doc_title_text, title_style))
    elements.append(Spacer(1, 6))

    headers = ["학번", "이름", "구분", "이수학년", "세부", "수정전 (원문)", "수정 후 (제안)", "수정 사유/근거", "수정구분"]
    table_data = [[Paragraph(h, header_cell_style) for h in headers]]
    col_widths = [40, 42, 35, 45, 50, 166.6, 166.6, 166.65, 45]

    for idx, row in audit_df.iterrows():
        r_data = [
            Paragraph(str(row.get('학번(숫자5자리)', row.get('학번', ''))), cell_style),
            Paragraph(str(row.get('이름', '')), cell_style),
            Paragraph(str(row.get('구분', '')), cell_style),
            Paragraph(str(row.get('이수학년', '')), cell_style),
            Paragraph(str(row.get('세부', '')), cell_style),
            Paragraph(str(row.get('수정전', '')), cell_style),
            Paragraph(str(row.get('수정 후', '')), cell_style),
            Paragraph(str(row.get('수정해야하는 이유나 근거', '')), cell_style),
            Paragraph(str(row.get('수정구분', '')), cell_style),
        ]
        table_data.append(r_data)

    t_style_cmds = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1E3A8A")),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
    ]

    for row_idx, row in enumerate(audit_df.to_dict('records'), start=1):
        sev_val = str(row.get('수정구분', '')).strip()
        if '필수' in sev_val:
            t_style_cmds.append(('BACKGROUND', (8, row_idx), (8, row_idx), colors.HexColor("#FEF08A")))

    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle(t_style_cmds))

    elements.append(t)
    doc.build(elements, canvasmaker=NumberedCanvas)
    return buffer.getvalue()


def create_refined_original_pdf_bytes(data_store: dict) -> bytes:
    """
    페이지 나눔/인쇄 경계 결합이 완료된 정제 원본 데이터(세특, 창체, 행특)를
    ReportLab 기반의 A4 가로 인쇄형 PDF 파일로 변환하여 바이트로 반환합니다.
    """
    font_name = register_korean_font()
    buffer = io.BytesIO()
    margin_pt = 42.52

    available_keys = [k for k in ['창체', '세특', '행특'] if data_store.get(k) is not None]
    cat_str = " · ".join(available_keys) if available_keys else "원본"
    
    file_names = data_store.get('file_names', {})
    base_names = [os.path.splitext(file_names[t])[0] for t in available_keys if t in file_names]
    sub_title_file = f" ({', '.join(base_names)})" if base_names else ""

    doc_title_text = f"학교생활기록부 정제 원본 리포트 - {cat_str}{sub_title_file}"

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=margin_pt,
        rightMargin=margin_pt,
        topMargin=margin_pt,
        bottomMargin=margin_pt + 15
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'PDFDocTitleRefined',
        parent=styles['Heading1'],
        fontName=font_name,
        fontSize=13,
        leading=16,
        textColor=colors.HexColor("#0F172A"),
        spaceAfter=8
    )
    cat_header_style = ParagraphStyle(
        'PDFCatHeaderStyle',
        parent=styles['Heading2'],
        fontName=font_name,
        fontSize=11,
        leading=14,
        textColor=colors.HexColor("#1E3A8A"),
        spaceBefore=10,
        spaceAfter=6
    )
    cell_style = ParagraphStyle(
        'PDFCellTextRefined',
        fontName=font_name,
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#1F2937")
    )
    header_cell_style = ParagraphStyle(
        'PDFHeaderCellTextRefined',
        fontName=font_name,
        fontSize=8,
        leading=10,
        textColor=colors.white,
        alignment=1
    )

    elements = []
    elements.append(Paragraph(doc_title_text, title_style))
    elements.append(Spacer(1, 4))

    for key in ['창체', '세특', '행특']:
        if key not in data_store or data_store[key] is None:
            continue
        
        item = data_store[key]
        if not isinstance(item, dict) or 'df' not in item:
            continue
        df = item['df']
        c_map = item['col_map']
        if df is None or df.empty:
            continue

        elements.append(Paragraph(f"■ [{key}] 정제 완료 원본 데이터 (총 {len(df)}건)", cat_header_style))
        elements.append(Spacer(1, 4))

        num_c = c_map['num_col']
        name_c = c_map['name_col']
        content_c = c_map['content_col']
        area_c = c_map.get('area_col')

        headers = ["학번", "이름", "구분/과목", "정제 결합된 원문 내용", "글자수"]
        col_widths = [45, 45, 90, 530, 46.85]

        table_data = [[Paragraph(h, header_cell_style) for h in headers]]

        for _, r in df.iterrows():
            num_val = str(r.get(num_c, '')).strip()
            name_val = str(r.get(name_c, '')).strip()
            
            if key == "세특" and '과목명' in r:
                sub_val = str(r['과목명']).strip()
            elif key == "창체" and '세부' in r:
                sub_val = str(r['세부']).strip()
            elif area_c and area_c in r and pd.notna(r[area_c]):
                sub_val = str(r[area_c]).strip()
            else:
                sub_val = key

            content_text = str(r.get(content_c, '')).strip()
            char_len = len(content_text)

            r_data = [
                Paragraph(num_val, cell_style),
                Paragraph(name_val, cell_style),
                Paragraph(sub_val, cell_style),
                Paragraph(content_text, cell_style),
                Paragraph(f"{char_len}자", cell_style),
            ]
            table_data.append(r_data)

        t = Table(table_data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1E3A8A")),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 12))

    doc.build(elements, canvasmaker=NumberedCanvas)
    return buffer.getvalue()



def create_formatted_excel_bytes(data_dict: dict) -> bytes:
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
    body_font = Font(name="맑은 고딕", size=9.5)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    for sheet_name, df in data_dict.items():
        if df is None or df.empty:
            continue
        
        export_df = df.copy()
        drop_cols = [c for c in export_df.columns if str(c).startswith('_')]
        export_df = export_df.drop(columns=drop_cols)

        ws = wb.create_sheet(title=sheet_name)
        headers = list(export_df.columns)
        ws.append(headers)

        ws.row_dimensions[1].height = 28
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

        for row in export_df.itertuples(index=False):
            ws.append(list(row))

        for r_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[r_idx].height = 42
            for c_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.font = body_font
                cell.border = thin_border
                col_name = str(headers[c_idx - 1])
                if col_name in ['번호', '성명', '이름', '과목명', '영역', '글자수', '학기']:
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align

        for col_idx, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)
            header_str = str(header)
            if '내용' in header_str or '특기사항' in header_str or '의견' in header_str:
                ws.column_dimensions[col_letter].width = 65
            elif '과목' in header_str or '영역' in header_str:
                ws.column_dimensions[col_letter].width = 18
            elif '성명' in header_str or '이름' in header_str:
                ws.column_dimensions[col_letter].width = 12
            elif '번호' in header_str or '글자수' in header_str:
                ws.column_dimensions[col_letter].width = 10
            else:
                ws.column_dimensions[col_letter].width = 16

    wb.save(output)
    return output.getvalue()


# ==============================================================================
# 메인 웹앱 라이프사이클 (Streamlit App Layout)
# ==============================================================================
def main():
    if 'data_store' not in st.session_state:
        st.session_state['data_store'] = {
            '창체': None,
            '세특': None,
            '행특': None,
            'raw_data': {},
            'merge_logs': {}
        }
    if 'llm_audit_results' not in st.session_state:
        st.session_state['llm_audit_results'] = None
    if 'has_audited' not in st.session_state:
        st.session_state['has_audited'] = False
    if 'uploader_key_version' not in st.session_state:
        st.session_state['uploader_key_version'] = 0

    if 'api_key_store' not in st.session_state:
        st.session_state['api_key_store'] = {
            "Gemini": os.getenv("GEMINI_API_KEY", ""),
            "OpenAI": os.getenv("OPENAI_API_KEY", ""),
            "Claude": os.getenv("CLAUDE_API_KEY", "")
        }

    guideline_text = load_guideline_content()

    # --------------------------------------------------------------------------
    # Global CSS injection (Direct DOM injection via st.markdown)
    # --------------------------------------------------------------------------
    st.markdown("""
        <style>
            /* 1. Real Dropzone Box Styling */
            section[role="presentation"],
            [data-testid="stFileUploaderDropzone"] {
                background-color: #F8FAFC !important;
                border: 2px dashed #94A3B8 !important;
                border-radius: 12px !important;
                padding: 1.4rem 1rem !important;
                display: flex !important;
                flex-direction: column !important;
                align-items: center !important;
                justify-content: center !important;
                text-align: center !important;
                cursor: pointer !important;
                min-height: 85px !important;
                transition: all 0.2s ease !important;
            }

            section[role="presentation"]:hover,
            [data-testid="stFileUploaderDropzone"]:hover {
                border-color: #2563EB !important;
                background-color: #EFF6FF !important;
            }

            /* Hide ALL internal Upload buttons and small helper text */
            section[role="presentation"] button,
            section[role="presentation"] > span,
            section[role="presentation"] small,
            [data-testid="stFileUploaderDropzone"] button,
            [data-testid="stFileUploaderDropzone"] small {
                display: none !important;
                visibility: hidden !important;
                height: 0 !important;
                width: 0 !important;
                overflow: hidden !important;
                position: absolute !important;
            }

            /* Clean Dropzone Label Text */
            section[role="presentation"]::after,
            [data-testid="stFileUploaderDropzone"]::after {
                content: "여기에 생기부 엑셀 파일 드롭\\A(또는 클릭하여 파일 선택)" !important;
                white-space: pre-wrap !important;
                font-size: 0.9rem !important;
                font-weight: 700 !important;
                color: #1E293B !important;
                line-height: 1.6 !important;
                text-align: center !important;
                display: block !important;
                pointer-events: none !important;
            }

            /* -------------------------------------------------------------
               ABSOLUTELY ZERO GAP BETWEEN ALL SIDEBAR ELEMENTS
               ------------------------------------------------------------- */
            [data-testid="stSidebar"] [data-testid="stVerticalBlock"] {
                gap: 0px !important;
                row-gap: 0px !important;
            }

            [data-testid="stSidebar"] [data-testid="stElementContainer"] {
                margin: 0px !important;
                margin-top: 0px !important;
                margin-bottom: 0px !important;
                padding: 0px !important;
                padding-top: 0px !important;
                padding-bottom: 0px !important;
            }

            [data-testid="stSidebar"] [data-testid="stExpander"],
            [data-testid="stSidebar"] details {
                margin: 0px !important;
                margin-top: 0px !important;
                margin-bottom: 0px !important;
            }

            [data-testid="stSidebar"] hr,
            [data-testid="stSidebar"] [data-testid="stMarkdownContainer"] hr {
                margin: 0px !important;
                margin-top: 0px !important;
                margin-bottom: 0px !important;
                padding: 0px !important;
                border: none !important;
                border-top: 1.5px solid #94A3B8 !important;
            }

            [data-testid="stSidebar"] .stMarkdown p,
            [data-testid="stSidebar"] .stMarkdown div,
            [data-testid="stSidebar"] [data-testid="stMarkdownContainer"] p,
            [data-testid="stSidebar"] [data-testid="stMarkdownContainer"] div {
                margin-top: 0px !important;
                margin-bottom: 0px !important;
            }

            [data-testid="stSidebar"] [data-testid="stButton"],
            [data-testid="stSidebar"] [data-testid="stBaseButton-secondary"] {
                margin: 0px !important;
                margin-top: 0px !important;
                margin-bottom: 0px !important;
            }

            /* Expander internals */
            [data-testid="stSidebar"] [data-testid="stExpander"] [data-testid="stExpanderDetails"] {
                padding: 0.65rem !important;
            }
            [data-testid="stSidebar"] [data-testid="stExpander"] [data-testid="stVerticalBlock"] {
                gap: 0.65rem !important;
                row-gap: 0.65rem !important;
            }
        </style>
    """, unsafe_allow_html=True)

    # --------------------------------------------------------------------------
    # 사이드바 Layout & Expanders (정확히 8개의 메인 요소로 구성)
    # --------------------------------------------------------------------------
    with st.sidebar:
        # 1. 제목 및 부제목 통합 블록
        st.markdown("""
            <div style="text-align: center; margin: 0; padding: 1.0rem 0 1.0rem 0;">
                <h2 style="font-size: 1.45rem; font-weight: 800; color: #0F172A; margin: 0; padding: 0; line-height: 1.2; letter-spacing: -0.4px;">꼼체크체크</h2>
                <div style="font-size: 0.88rem; font-weight: 600; color: #475569; padding-top: 0.5rem !important; padding-bottom: 1.5rem !important; margin: 0 !important; letter-spacing: -0.2px;">생기부 AI 검증 시스템</div>
            </div>
        """, unsafe_allow_html=True)
        
        # 2. 상단 구분선
        st.markdown("---")

        # 3. AI API 설정 (Google Gemini 전용)
        st.markdown("<div style='height: 1.5rem;'></div>", unsafe_allow_html=True)
        with st.expander("AI API 설정 (Gemini)", expanded=False):
            provider = "Gemini"
            key_input_id = "api_key_input_Gemini"
            if key_input_id not in st.session_state:
                st.session_state[key_input_id] = st.session_state['api_key_store'].get("Gemini", "")

            st.markdown('<div style="color: #B91C1C; font-weight: 700; font-size: 0.88rem; margin-bottom: 0.3rem;">Gemini API Key 입력</div>', unsafe_allow_html=True)
            api_key = st.text_input(
                "Gemini API Key 입력",
                label_visibility="collapsed",
                type="password",
                key=key_input_id,
                help="Google AI Studio에서 발급받은 개인 Gemini API Key를 입력해 주세요."
            )
            st.session_state['api_key_store']["Gemini"] = api_key
            
            # API 키로 Google AI Studio에서 지원 가능한 Gemini 모델 실시간 동적 조회
            fetched_models = fetch_available_gemini_models(api_key) if api_key else []
            
            gemini_model_api_map = {}
            if fetched_models:
                for fm in fetched_models:
                    gemini_model_api_map[fm["display"]] = fm["id"]
                
                gemini_model_options = list(gemini_model_api_map.keys())
                default_idx = 0
                for i, opt in enumerate(gemini_model_options):
                    if "2.5-flash" in opt.lower():
                        default_idx = i
                        break
                    elif "3.1-pro" in opt.lower():
                        default_idx = i
                    elif "1.5-pro" in opt.lower():
                        default_idx = i
            else:
                # API 키 입력 전 또는 기본 정적 선택지
                gemini_model_api_map = {
                    "Gemini 2.5 Flash (gemini-2.5-flash - 실시간 검증 완료/초고속) 🚀🌟": "gemini-2.5-flash",
                    "Gemini 3.1 Pro Preview (gemini-3.1-pro-preview - 최신 차세대 Pro) 🌟": "gemini-3.1-pro-preview",
                    "Gemini 1.5 Pro (gemini-1.5-pro - 검증된 최고 안정 Pro)": "gemini-1.5-pro",
                    "Gemini 2.0 Flash (gemini-2.0-flash - 초고속 Flash) ⚡": "gemini-2.0-flash",
                    "Gemini 1.5 Flash (gemini-1.5-flash - 가장 안정적) ⚡": "gemini-1.5-flash",
                    "Gemini 1.0 Pro (gemini-1.0-pro - 구버전 Pro)": "gemini-1.0-pro",
                    "Gemini Flash Latest (gemini-flash-latest)": "gemini-flash-latest",
                    "Gemini Pro Latest (gemini-pro-latest)": "gemini-pro-latest"
                }
                gemini_model_options = list(gemini_model_api_map.keys())
                default_idx = 0

            gemini_model_selected = st.selectbox(
                "Gemini 모델 선택",
                gemini_model_options,
                index=default_idx,
                help="Google AI Studio API에서 제공하는 텍스트 생성 Gemini 모델 목록입니다."
            )
            model_name = gemini_model_api_map.get(gemini_model_selected, "gemini-2.5-flash")

            if st.button("API Key 삭제", key="btn_del_key_Gemini", help="입력한 Gemini API Key만 즉시 삭제하고 초기화합니다.", use_container_width=True):
                st.session_state['api_key_store']["Gemini"] = ""
                st.session_state[key_input_id] = ""
                st.rerun()

            with st.expander("Gemini API Key 발급 가이드"):
                st.markdown("""
                **Google Gemini API Key 발급 방법**
                1. [Google AI Studio](https://aistudio.google.com/) 접속 후 로그인
                2. **Dashboard -> API 키 -> Create API key** 클릭
                3. 생성된 키를 위의 입력란에 붙여넣기
                """)

        # 4. 엑셀 파일 업로드
        with st.expander("엑셀 파일 업로드", expanded=True):
            uploader_key = f"auto_file_uploader_{st.session_state.get('uploader_key_version', 0)}"
            uploaded_files = st.file_uploader(
                "생기부 엑셀 파일 (.xlsx, .xls)",
                type=["xlsx", "xls"],
                accept_multiple_files=True,
                key=uploader_key,
                label_visibility="collapsed"
            )

            # 파일이 비어있는 상태에서 직전에 2개 이상 파일 업로드 시도가 있었다면 경고 배너 출력
            if not uploaded_files and st.session_state.get('multi_file_warning', False):
                st.markdown("""
                    <div style="background-color: #FEF2F2; border: 1px solid #FCA5A5; color: #991B1B; border-radius: 8px; padding: 0.55rem 0.6rem; font-size: 0.82rem; text-align: center; margin: 0.4rem 0 0 0; line-height: 1.45; width: 100%; box-sizing: border-box; font-weight: 600;">
                        ⚠️ 한 번에 1개의 파일만 업로드할 수 있습니다.<br>
                        <span style="font-size: 0.76rem; font-weight: 500; color: #B91C1C;">(2개 이상의 파일이 감지되어 취소되었습니다 ➔ 1개 파일만 드래그해 주세요)</span>
                    </div>
                """, unsafe_allow_html=True)

            if uploaded_files:
                if len(uploaded_files) > 1:
                    # 2개 이상 들어오면 즉시 업로더를 리셋하여 드롭존에서 완전히 퇴출
                    st.session_state['multi_file_warning'] = True
                    st.session_state['uploader_key_version'] = st.session_state.get('uploader_key_version', 0) + 1
                    st.rerun()
                else:
                    # 정상적으로 1개 파일만 들어온 경우 경고 해제
                    st.session_state['multi_file_warning'] = False
                    uploaded_file = uploaded_files[0]
                    try:
                        try:
                            raw_df = pd.read_excel(uploaded_file, header=None)
                        except Exception:
                            uploaded_file.seek(0)
                            html_dfs = pd.read_html(uploaded_file)
                            if html_dfs:
                                raw_df = html_dfs[0]
                            else:
                                raise
                        type_key = detect_actual_record_type(uploaded_file.name, raw_df)

                        file_sig = f"{type_key}_{uploaded_file.name}_{uploaded_file.size}"
                        if 'file_signatures' not in st.session_state:
                            st.session_state['file_signatures'] = {}

                        if st.session_state['file_signatures'].get(type_key) != file_sig:
                            if st.session_state.get('llm_audit_results') is not None or st.session_state.get('has_audited', False):
                                st.session_state['data_store'] = {'raw_data': {}, 'merge_logs': {}}
                                st.session_state['llm_audit_results'] = None
                                st.session_state['has_audited'] = False
                                st.info("새 파일 업로드가 감지되어 이전 분석 결과 및 기존 데이터가 자동 초기화되었습니다.")

                            st.session_state['file_signatures'][type_key] = file_sig

                        st.session_state['data_store']['raw_data'][type_key] = raw_df
                        if 'file_names' not in st.session_state['data_store']:
                            st.session_state['data_store']['file_names'] = {}
                        st.session_state['data_store']['file_names'][type_key] = uploaded_file.name

                        df_processed, col_map = detect_columns(raw_df)
                        refined_df, logs, hope_blanks = refine_student_records(df_processed, col_map)
                        
                        if 'hope_blank_records' not in st.session_state['data_store']:
                            st.session_state['data_store']['hope_blank_records'] = []
                        st.session_state['data_store']['hope_blank_records'].extend(hope_blanks)

                        if type_key == "세특":
                            final_df = split_subject_details(refined_df, col_map)
                        else:
                            final_df = refined_df.copy()
                            content_c = col_map['content_col']
                            final_df['글자수'] = get_safe_series(final_df, content_c).astype(str).apply(len)

                        st.session_state['data_store'][type_key] = {
                            'df': final_df,
                            'raw_df': raw_df,
                            'col_map': col_map
                        }
                        st.session_state['data_store']['merge_logs'][type_key] = logs

                        num_c, name_c = col_map['num_col'], col_map['name_col']
                        num_series = get_safe_series(final_df, num_c)
                        name_series = get_safe_series(final_df, name_c)
                        unique_students_cnt = len(pd.DataFrame({'num': num_series, 'name': name_series}).drop_duplicates())
                        
                        if type_key == "창체":
                            actual_records_cnt = len(final_df)
                            hb_list = st.session_state.get('data_store', {}).get('hope_blank_records', [])
                            blank_cnt = len(hb_list) if hb_list else 0

                            # 학년 구성 자동 분석 (단일 학년 vs 다개년 누적)
                            grade_series = get_safe_series(final_df, col_map.get('grade_col', ''))
                            unique_grades = sorted([str(g).strip() for g in grade_series.dropna().unique() if str(g).strip() not in ['', 'nan', 'None']])
                            is_multi_grade = len(unique_grades) > 1

                            if is_multi_grade:
                                grades_label = f" [{', '.join(unique_grades)} 누적]"
                                detect_msg = f"<b>[{type_key}] 감지 완료!</b> (총 {unique_students_cnt}명 학생{grades_label}, {actual_records_cnt}개 영역 100% 감지 완료)"
                            else:
                                expected_total = unique_students_cnt * 3
                                if actual_records_cnt == expected_total:
                                    detect_msg = f"<b>[{type_key}] 감지 완료!</b> (총 {unique_students_cnt}명 학생, {expected_total}개 영역 100% 감지 완료)"
                                elif actual_records_cnt + blank_cnt == expected_total:
                                    detect_msg = f"<b>[{type_key}] 감지 완료!</b> (총 {unique_students_cnt}명 학생, 서술문 {actual_records_cnt}개 + 희망분야 공란 {blank_cnt}개 100% 감지)"
                                else:
                                    missing_cnt = expected_total - actual_records_cnt
                                    if missing_cnt > 0:
                                        detect_msg = f"<b>[{type_key}] 감지 완료!</b> (총 {unique_students_cnt}명 학생, {actual_records_cnt}개 영역 기록, 미작성 {missing_cnt}개 영역)"
                                    else:
                                        detect_msg = f"<b>[{type_key}] 감지 완료!</b> (총 {unique_students_cnt}명 학생, {actual_records_cnt}개 영역 기록)"
                        elif type_key == "세특":
                            detect_msg = f"<b>[{type_key}] 감지 완료!</b> (총 {unique_students_cnt}명 학생, {len(final_df)}개 과목 기록)"
                        else:
                            detect_msg = f"<b>[{type_key}] 감지 완료!</b> (총 {unique_students_cnt}명 학생)"

                        st.markdown(f"""
                            <div style="background-color: #ECFDF5; border: 1px solid #A7F3D0; color: #065F46; border-radius: 8px; padding: 0.45rem 0.5rem; font-size: 0.83rem; text-align: center; margin: 0; line-height: 1.35; width: 100%; box-sizing: border-box;">
                                ✅ {detect_msg}
                            </div>
                        """, unsafe_allow_html=True)

                    except Exception as e:
                        st.markdown(f"""
                            <div style="background-color: #FEF2F2; border: 1px solid #FCA5A5; color: #991B1B; border-radius: 8px; padding: 0.2rem 0.2rem 0.5rem 0.5rem; font-size: 0.83rem; text-align: center; margin: 0; line-height: 1.35; width: 100%; box-sizing: border-box;">
                                ⚠️ 파일 처리 오류 ({uploaded_file.name}): {str(e)}
                            </div>
                        """, unsafe_allow_html=True)
                        with st.expander("오류 상세 내용"):
                            st.code(traceback.format_exc())

            if any(st.session_state['data_store'].get(k) is not None for k in ['창체', '세특', '행특']):
                btn_run_llm = st.button("AI 정밀 분석 실행", type="primary", use_container_width=True, key="sidebar_btn_run_llm")
                if btn_run_llm:
                    st.session_state['sidebar_manage_mode_radio'] = "메인 AI 정밀 검증"
            else:
                btn_run_llm = False

        # 5. 데이터 관리 메뉴
        with st.expander("데이터 관리 메뉴", expanded=False):
            manage_mode = st.radio(
                "메뉴 기능 선택",
                [
                    "메인 AI 정밀 검증",
                    "학생별 통합 조회", 
                    "페이지 나눔 정제 검증", 
                    "데이터 분석 & 통계"
                ],
                key="sidebar_manage_mode_radio"
            )

        # 6. 하단 구분선
        st.markdown("<div style='height: 0.5rem;'></div>", unsafe_allow_html=True)
        st.markdown("---")
        
        # 7. 분석 결과 초기화 버튼
        st.markdown("<div style='height: 1.5rem;'></div>", unsafe_allow_html=True)
        if st.button("분석 결과 초기화", use_container_width=True, help="업로드된 모든 생기부 파일과 AI 검증 결과를 초기화하고 새로 시작합니다."):
            st.session_state['uploader_key_version'] = st.session_state.get('uploader_key_version', 0) + 1
            st.session_state['data_store'] = {'raw_data': {}, 'merge_logs': {}}
            st.session_state['llm_audit_results'] = None
            st.session_state['has_audited'] = False
            st.session_state['file_signatures'] = {}
            st.rerun()

    available_types = [k for k in ['창체', '세특', '행특'] if st.session_state['data_store'].get(k) is not None]

    # --------------------------------------------------------------------------
    # MODE 1: 메인 AI 정밀 검증 & 리포트 다운로드
    # --------------------------------------------------------------------------
    if manage_mode == "메인 AI 정밀 검증":
        if not available_types:
            st.markdown("""
                <div style="text-align: center; padding: 2.8rem 1.5rem; background: #FFFFFF; border: 1px dashed #CBD5E1; border-radius: 16px; margin: 1rem 0;">
                    <h3 style="font-size: 1.15rem; font-weight: 700; color: #1E293B; margin-bottom: 0.4rem;">업로드된 생활기록부 엑셀 파일이 없습니다</h3>
                    <p style="font-size: 0.9rem; color: #64748B; margin-bottom: 1rem;">왼쪽 사이드바에서 세특, 창체, 또는 행특 엑셀 파일을 업로드해 주세요.</p>
                    <div style="display: inline-block; text-align: left; background: #F8FAFC; border: 1px solid #E2E8F0; border-radius: 10px; padding: 0.9rem 1.2rem; font-size: 0.85rem; color: #475569; max-width: 600px; line-height: 1.6;">
                        📌 <b>NEIS 엑셀 파일 다운로드 안내</b><br>
                        <b>나이스(NEIS) 접속 ➔ 학교생활기록부 ➔ 학생부 영역별 조회</b>에서 아래 영역을 <b>XLS 형식</b>으로 다운로드해 사용해 주세요. <span style="color: #DC2626; font-weight: 600;">(※XLS data 형식으로 다운로드 하면 안됩니다.)</span>
                        <ul style="margin: 0.4rem 0 0 1.2rem; padding: 0;">
                            <li><b>창의적체험활동</b></li>
                            <li><b>교과학습발달상황</b> (세부능력및특기사항)</li>
                            <li><b>행동특성및종합의견</b></li>
                        </ul>
                    </div>
                </div>
            """, unsafe_allow_html=True)
        else:
            if btn_run_llm:
                if not api_key:
                    st.error("AI API Key가 입력되지 않았습니다. 사이드바에서 API Key를 입력해 주세요.")
                else:
                    progress_container = st.container()
                    with progress_container:
                        progress_bar = st.progress(0.01)
                        status_text = st.empty()
                        status_text.info("AI 정밀 검사 시작 중... (데이터 패키징 및 AI 모델 연결 ⏳)")
                        
                        st.markdown("<div style='height: 0.5rem;'></div>", unsafe_allow_html=True)
                        st.markdown("##### 📋 실시간 학생별 검증 현황")
                        log_container = st.empty()
                        live_log_records = []

                        def update_progress(event_type, current_b, total_b, st_label="학생", batch_items=None):
                            if event_type == "start":
                                pct = (current_b) / total_b
                                progress_bar.progress(max(0.01, min(0.99, pct)))
                                status_text.info(f"AI 정밀 검사 진행 중... [{current_b + 1}/{total_b} 배치: **{st_label}** 분석 요청 중 ⏳]")
                                return

                            # event_type == "finish"
                            pct = current_b / total_b
                            progress_bar.progress(max(0.01, min(1.0, pct)))
                            status_text.info(f"AI 정밀 검사 진행 중... [{current_b}/{total_b} 배치 완료 ({st_label})]")
                            
                            items_list = batch_items if batch_items is not None else []
                            req_cnt = sum(1 for x in items_list if "필수" in str(x.get('severity', '')))
                            rec_cnt = sum(1 for x in items_list if "권장" in str(x.get('severity', '')) and "검토" not in str(x.get('severity', '')))
                            rev_cnt = sum(1 for x in items_list if "검토" in str(x.get('severity', '')))
                            
                            badge_parts = []
                            if req_cnt > 0:
                                badge_parts.append(f"<span style='background:#FEE2E2;color:#991B1B;padding:2px 8px;border-radius:6px;font-weight:700;font-size:0.8rem;'>🔴 수정 필수 {req_cnt}건</span>")
                            if rec_cnt > 0:
                                badge_parts.append(f"<span style='background:#FEF3C7;color:#92400E;padding:2px 8px;border-radius:6px;font-weight:700;font-size:0.8rem;'>🟡 수정 권장 {rec_cnt}건</span>")
                            if rev_cnt > 0:
                                badge_parts.append(f"<span style='background:#E0E7FF;color:#3730A3;padding:2px 8px;border-radius:6px;font-weight:700;font-size:0.8rem;'>🔵 검토 권장 {rev_cnt}건</span>")
                            if not badge_parts:
                                badge_parts.append("<span style='background:#ECFDF5;color:#065F46;padding:2px 8px;border-radius:6px;font-weight:700;font-size:0.8rem;'>✅ 이상 없음 (0건)</span>")
                            
                            badges_html = " ".join(badge_parts)
                            
                            items_summary = ""
                            if items_list:
                                sample_snippets = []
                                for itm in items_list[:3]:
                                    sub_s = itm.get('sub_category', '')
                                    orig_s = itm.get('original_text', '')
                                    sugg_s = itm.get('suggested_text', '')
                                    sev_s = itm.get('severity', '수정 필수')
                                    dot = "🔴" if "필수" in sev_s else ("🟡" if ("권장" in sev_s and "검토" not in sev_s) else "🔵")
                                    sample_snippets.append(f"<div style='font-size:0.82rem;color:#475569;margin-left:0.5rem;margin-top:0.2rem;'>• {dot} <b>[{sub_s}]</b> {orig_s} ➔ <b>{sugg_s}</b> <span style='color:#64748B;'>({itm.get('reason','')})</span></div>")
                                if len(items_list) > 3:
                                    sample_snippets.append(f"<div style='font-size:0.78rem;color:#94A3B8;margin-left:0.5rem;margin-top:0.2rem;'>... 외 {len(items_list)-3}건</div>")
                                items_summary = "".join(sample_snippets)
                            
                            card_html = (
                                f"<div style='background:#FFFFFF;border:1px solid #E2E8F0;border-radius:10px;padding:0.7rem 1rem;margin-bottom:0.6rem;box-shadow:0 1px 3px rgba(0,0,0,0.03);'>"
                                f"<div style='display:flex;justify-content:space-between;align-items:center;margin-bottom:0.2rem;'>"
                                f"<div style='font-weight:700;color:#1E293B;font-size:0.92rem;'>[{current_b}/{total_b}] {st_label}</div>"
                                f"<div>{badges_html}</div>"
                                f"</div>"
                                f"{items_summary}"
                                f"</div>"
                            )
                            live_log_records.insert(0, card_html)
                            all_cards_html = "".join(live_log_records)
                            log_container.markdown(f"<div style='max-height:420px;overflow-y:auto;padding-right:4px;'>{all_cards_html}</div>", unsafe_allow_html=True)

                        try:
                            records_payload = prepare_records_for_llm(st.session_state['data_store'])
                            raw_findings = call_llm_api_for_audit(provider, api_key, model_name, records_payload, guideline_text, progress_callback=update_progress)
                            progress_bar.progress(1.0)
                            status_text.success("AI 정밀 검증이 성공적으로 완료되었습니다!")
                            
                            audit_rows = []
                            for item in raw_findings:
                                cat_item = str(item.get("category", item.get("구분", "창체"))).strip()
                                sub_cat_item = str(item.get("sub_category", item.get("세부", ""))).strip()
                                orig_text_item = str(item.get("original_text", item.get("수정전", ""))).strip()
                                st_id = str(item.get("student_id", item.get("학번", "00000"))).strip()

                                matched_payload = None
                                item_taken_grade = str(item.get("taken_grade", item.get("이수학년", ""))).strip()

                                # 0. 이수시간 0 관련 오류인 경우 해당 학생의 0시간 레코드 우선 매칭
                                if orig_text_item in ["0", "0시간", "0 시간"] or "이수시간" in str(item.get("reason", "")):
                                    matched_payload = next((p for p in records_payload if p["학번"] == st_id and str(p.get("이수시간", "")).strip() == "0" and (not sub_cat_item or p["세부"] == sub_cat_item)), None)

                                # 1. Exact match with student_id + taken_grade + original_text
                                if not matched_payload and orig_text_item and st_id:
                                    if item_taken_grade:
                                        matched_payload = next((p for p in records_payload if p["학번"] == st_id and p.get("이수학년") == item_taken_grade and orig_text_item in p["기록텍스트"]), None)
                                    if not matched_payload:
                                        matched_payload = next((p for p in records_payload if p["학번"] == st_id and orig_text_item in p["기록텍스트"]), None)
                                
                                # 2. Match with student_name + original_text in text_content
                                if not matched_payload and orig_text_item and item.get("student_name"):
                                    matched_payload = next((p for p in records_payload if p["이름"] == item.get("student_name") and orig_text_item in p["기록텍스트"]), None)

                                # 3. Match with student_id + taken_grade + sub_category
                                if not matched_payload and sub_cat_item and sub_cat_item not in ["창체", "창의적체험활동", "세특", "세부능력및특기사항"]:
                                    if item_taken_grade:
                                        matched_payload = next((p for p in records_payload if p["학번"] == st_id and p.get("이수학년") == item_taken_grade and p["세부"] == sub_cat_item), None)
                                    if not matched_payload:
                                        matched_payload = next((p for p in records_payload if p["학번"] == st_id and p["세부"] == sub_cat_item), None)
                                
                                # 4. Match with original_text across all records
                                if not matched_payload and orig_text_item:
                                    matched_payload = next((p for p in records_payload if orig_text_item in p["기록텍스트"]), None)

                                # 엑셀 원본 ground-truth 메타데이터로 100% 강제 동기화 (LLM 학년/영역 오추론 방지)
                                if matched_payload:
                                    true_grade = matched_payload.get("이수학년", item.get("taken_grade", "1학년"))
                                    true_name = matched_payload.get("이름", item.get("student_name", ""))
                                    true_cat = matched_payload.get("구분", cat_item)
                                    true_sub = matched_payload.get("세부", sub_cat_item)
                                else:
                                    true_grade = item.get("taken_grade", item.get("이수학년", "1학년"))
                                    true_name = item.get("student_name", item.get("이름", ""))
                                    true_cat = cat_item
                                    true_sub = sub_cat_item

                                if true_cat in ["창체", "창의적체험활동"] and (not true_sub or true_sub in ["창체", "창의적체험활동"]):
                                    true_sub = get_changche_area_from_cell(true_sub, orig_text_item)
                                elif true_cat == "세특" and true_sub in ["세부능력및특기사항", "세부능력 및 특기사항", "세특", "세부", "nan", ""]:
                                    true_sub = "개세특"

                                # 1. 분할된 불필요한 음절/글자 제거 관련 환각 필터링
                                reason_str = str(item.get("reason", item.get("수정해야하는 이유나 근거", "")))
                                if any(k in reason_str for k in ["분할된", "불필요한 음절", "음절을 제거", "글자를 제거", "불필요한 글자", "음절 제거", "글자 제거", "음절 삭제", "글자 삭제"]):
                                    continue

                                # 2. 학기 표기('(1학기)', '(2학기)') 삭제/제거 환각 필터링
                                suggested_str = str(item.get("suggested_text", item.get("수정 후", ""))).strip()
                                if any(k in reason_str for k in ["학기 정보", "학기 표기", "과목명만 기재", "학기를 제거", "학기 제거", "학기 정보는"]):
                                    continue
                                if re.match(r'^\([12]학기\)\s*', orig_text_item) and not re.match(r'^\([12]학기\)\s*', suggested_str):
                                    continue

                                # 3. 분할된 문장 잘림/미완결에 대한 임의 단어 보완 및 문장 완성 환각 필터링
                                if any(k in reason_str for k in [
                                    "중간에 잘려", "단어를 보완", "문장을 완성", "잘린 '", "잘린 \"",
                                    "결합하여 자연스러운", "결합하여 문맥", "의미가 연결되도록", "문맥으로 수정",
                                    "불완전한 문장", "문장 미완결", "문장이 종결되지", "문장이 끊겨", "단어가 잘려",
                                    "단어가 불완전", "앞선 문장에서 잘린"
                                ]):
                                    continue

                                audit_rows.append({
                                    "학번": st_id,
                                    "이름": true_name,
                                    "구분": true_cat,
                                    "이수학년": true_grade,
                                    "세부": true_sub,
                                    "수정전": orig_text_item,
                                    "수정 후": item.get("suggested_text", item.get("수정 후", "")),
                                    "수정해야하는 이유나 근거": item.get("reason", item.get("수정해야하는 이유나 근거", "")),
                                    "수정구분": item.get("severity", item.get("수정구분", "수정 필수"))
                                })

                            if audit_rows:
                                # 완전 동일 행 중복 제거 (Deduplication)
                                dedup_audit_rows = []
                                seen_keys = set()
                                for r in audit_rows:
                                    dedup_key = (
                                        str(r.get("학번", "")),
                                        str(r.get("이수학년", "")),
                                        str(r.get("세부", "")),
                                        str(r.get("수정전", "")).strip(),
                                        str(r.get("수정 후", "")).strip()
                                    )
                                    if dedup_key not in seen_keys:
                                        seen_keys.add(dedup_key)
                                        dedup_audit_rows.append(r)
                                audit_rows = dedup_audit_rows

                                res_df = pd.DataFrame(audit_rows)
                                sev_rank_map = {
                                    "수정 필수": 1, "수정필수": 1,
                                    "수정 권장": 2, "수정 권고": 2, "수정권장": 2, "수정권고": 2,
                                    "검토 권장": 3, "검토 권고": 3, "검토권장": 3, "검토권고": 3
                                }
                                area_order_map = {"자율활동": 1, "동아리활동": 2, "진로활동": 3}
                                res_df['_sev_rank'] = res_df['수정구분'].map(lambda x: sev_rank_map.get(str(x).strip(), 4))
                                res_df['_area_rank'] = res_df['세부'].map(lambda x: area_order_map.get(str(x).strip(), 4))
                                res_df = res_df.sort_values(by=['학번', '_area_rank', '_sev_rank', '이름'], ascending=[True, True, True, True]).reset_index(drop=True)
                                res_df = res_df.drop(columns=['_sev_rank', '_area_rank'])
                                st.session_state['llm_audit_results'] = res_df
                            else:
                                st.session_state['llm_audit_results'] = pd.DataFrame(columns=[
                                    "학번", "이름", "구분", "이수학년", "세부", "수정전", "수정 후", "수정해야하는 이유나 근거", "수정구분"
                                ])
                            st.session_state['has_audited'] = True

                            st.success("AI 정밀 검사 완료!")

                        except Exception as e:
                            st.error(f"AI 분석 중 오류가 발생했습니다: {str(e)}")
                            with st.expander("상세 에러 내역"):
                                st.code(traceback.format_exc())

            audit_df = st.session_state.get('llm_audit_results')

            if audit_df is not None and st.session_state.get('has_audited', False):
                req_cnt = len(audit_df[audit_df['수정구분'].str.contains('필수', na=False)]) if not audit_df.empty else 0
                rec_cnt = len(audit_df[audit_df['수정구분'].str.contains('수정 권장|수정권장|수정 권고|수정권고', na=False)]) if not audit_df.empty else 0
                rev_cnt = len(audit_df[audit_df['수정구분'].str.contains('검토', na=False)]) if not audit_df.empty else 0

                col_m1, col_m2, col_m3, col_m4, col_m5 = st.columns(5)
                with col_m1:
                    st.markdown(f'<div class="metric-card"><div class="val" style="color:#DC2626;">{len(audit_df)}</div><div class="lbl">총 검출 오류 건수</div></div>', unsafe_allow_html=True)
                with col_m2:
                    st.markdown(f'<div class="metric-card"><div class="val" style="color:#EF4444;">{req_cnt}</div><div class="lbl">수정 필수 (지침 위반/오타)</div></div>', unsafe_allow_html=True)
                with col_m3:
                    st.markdown(f'<div class="metric-card"><div class="val" style="color:#F59E0B;">{rec_cnt}</div><div class="lbl">수정 권장 (문맥/기재권고)</div></div>', unsafe_allow_html=True)
                with col_m4:
                    st.markdown(f'<div class="metric-card"><div class="val" style="color:#6366F1;">{rev_cnt}</div><div class="lbl">검토 권장 (학생관점/표현)</div></div>', unsafe_allow_html=True)
                with col_m5:
                    st.markdown(f'<div class="metric-card"><div class="val" style="color:#10B981;">{provider} AI</div><div class="lbl">사용된 분석 엔진</div></div>', unsafe_allow_html=True)

                st.markdown("---")

                if audit_df.empty:
                    st.balloons()
                    st.success("AI 검사 결과, 검출된 오탈자나 기재 지침 위반 항목이 없습니다.")
                else:
                    col_f1, col_f2 = st.columns([2, 3])
                    with col_f1:
                        filter_cat = st.selectbox("수정 구분 필터", ["전체 보기", "수정 필수만 보기", "수정 권장만 보기", "검토 권장만 보기"])
                    with col_f2:
                        search_keyword = st.text_input("학생 이름/학번 검색", placeholder="예: 10101 또는 김철수")

                    filtered_df = audit_df.copy()
                    if filter_cat == "수정 필수만 보기":
                        filtered_df = filtered_df[filtered_df["수정구분"].str.contains('필수', na=False)]
                    elif filter_cat == "수정 권장만 보기":
                        filtered_df = filtered_df[filtered_df["수정구분"].str.contains('수정 권장|수정권장|수정 권고|수정권고', na=False)]
                    elif filter_cat == "검토 권장만 보기":
                        filtered_df = filtered_df[filtered_df["수정구분"].str.contains('검토', na=False)]

                    if search_keyword.strip():
                        kw = search_keyword.strip()
                        filtered_df = filtered_df[
                            filtered_df["학번"].astype(str).str.contains(kw) | filtered_df["이름"].astype(str).str.contains(kw)
                        ]

                    filtered_df.index = range(1, len(filtered_df) + 1)
                    st.markdown("### AI 오탈자 및 검증 결과 표")
                    st.dataframe(filtered_df, use_container_width=True, height=420)

                    st.markdown("---")
                    st.markdown("### 검증 리포트 & 데이터 다운로드")

                    req_df_out = filtered_df[filtered_df['수정구분'].str.contains('필수', na=False)]
                    # 사용자 지침: '수정 권장'용 분리 파일에는 '검토 권장'도 함께 표시
                    rec_df_out = filtered_df[filtered_df['수정구분'].str.contains('권장|권고|검토', na=False) & ~filtered_df['수정구분'].str.contains('필수', na=False)]

                    file_names = st.session_state['data_store'].get('file_names', {})
                    base_names = [os.path.splitext(file_names[t])[0] for t in available_types if t in file_names]
                    prefix = f"생기부_{'_'.join(base_names)}" if base_names else "생기부"

                    c_dl1, c_dl2 = st.columns(2)

                    with c_dl1:
                        st.markdown(f"#### 수정 필수 ({len(req_df_out)}건)")
                        b1, b2 = st.columns(2)
                        with b1:
                            st.download_button(
                                "엑셀 (.xlsx)",
                                data=create_audit_report_excel_bytes(req_df_out),
                                file_name=f"{prefix}_수정필수.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="dl_req_excel",
                                use_container_width=True,
                                disabled=req_df_out.empty
                            )
                        with b2:
                            st.download_button(
                                "PDF (.pdf)",
                                data=create_audit_report_pdf_bytes(req_df_out),
                                file_name=f"{prefix}_수정필수.pdf",
                                mime="application/pdf",
                                key="dl_req_pdf",
                                use_container_width=True,
                                disabled=req_df_out.empty
                            )

                    with c_dl2:
                        st.markdown(f"#### 수정 권장 (검토 권장 포함) ({len(rec_df_out)}건)")
                        b3, b4 = st.columns(2)
                        with b3:
                            st.download_button(
                                "엑셀 (.xlsx)",
                                data=create_audit_report_excel_bytes(rec_df_out),
                                file_name=f"{prefix}_수정권장.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="dl_rec_excel",
                                use_container_width=True,
                                disabled=rec_df_out.empty
                            )
                        with b4:
                            st.download_button(
                                "PDF (.pdf)",
                                data=create_audit_report_pdf_bytes(rec_df_out),
                                file_name=f"{prefix}_수정권장.pdf",
                                mime="application/pdf",
                                key="dl_rec_pdf",
                                use_container_width=True,
                                disabled=rec_df_out.empty
                            )

                    st.markdown("<div style='margin-top: 0.6rem;'></div>", unsafe_allow_html=True)
                    c_dl3, c_dl4 = st.columns(2)

                    with c_dl3:
                        st.markdown(f"#### 전체 통합 검증 ({len(filtered_df)}건)")
                        b_all1, b_all2 = st.columns(2)
                        with b_all1:
                            st.download_button(
                                "엑셀 (.xlsx)",
                                data=create_audit_report_excel_bytes(filtered_df),
                                file_name=f"{prefix}_전체리포트.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="dl_all_excel",
                                use_container_width=True,
                                disabled=filtered_df.empty
                            )
                        with b_all2:
                            st.download_button(
                                "PDF (.pdf)",
                                data=create_audit_report_pdf_bytes(filtered_df),
                                file_name=f"{prefix}_전체리포트.pdf",
                                mime="application/pdf",
                                key="dl_all_pdf",
                                use_container_width=True,
                                disabled=filtered_df.empty
                            )

                    with c_dl4:
                        available_exports = {t: st.session_state['data_store'][t]['df'] for t in available_types if t in st.session_state['data_store']}
                        st.markdown(f"#### 정제 원본 데이터 ({', '.join(available_exports.keys())})")
                        b_clean1, b_clean2 = st.columns(2)
                        with b_clean1:
                            st.download_button(
                                "엑셀 (.xlsx)",
                                data=create_formatted_excel_bytes(available_exports),
                                file_name=f"{prefix}_정제원본.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="dl_clean_excel",
                                use_container_width=True,
                                disabled=not available_exports
                            )
                        with b_clean2:
                            st.download_button(
                                "PDF (.pdf)",
                                data=create_refined_original_pdf_bytes(st.session_state['data_store']),
                                file_name=f"{prefix}_정제원본.pdf",
                                mime="application/pdf",
                                key="dl_clean_pdf",
                                use_container_width=True,
                                disabled=not available_exports
                            )

    # --------------------------------------------------------------------------
    # MODE 2: 학생별 통합 조회 (관리 메뉴 모드)
    # --------------------------------------------------------------------------
    elif manage_mode == "학생별 통합 조회":
        st.subheader("학생별 생기부 기록 통합 뷰어")

        if not available_types:
            st.warning("먼저 사이드바에서 생기부 엑셀 파일을 업로드해 주세요.")
        else:
            student_set = set()
            for t_key in available_types:
                data_item = st.session_state['data_store'][t_key]
                df_temp = data_item['df']
                c_map = data_item['col_map']
                num_c, name_c = c_map['num_col'], c_map['name_col']
                for _, r in df_temp.iterrows():
                    num_val, name_val = safe_val(r.get(num_c, '')), safe_val(r.get(name_c, ''))
                    if safe_notna(num_val) and safe_notna(name_val) and safe_str(name_val) != "":
                        student_set.add((safe_str(num_val), safe_str(name_val)))

            sorted_students = sorted(list(student_set), key=lambda x: int(re.sub(r'\D', '', x[0])) if re.sub(r'\D', '', x[0]).isdigit() else 999)

            if sorted_students:
                col_sel1, col_sel2 = st.columns([2, 3])
                with col_sel1:
                    selected_student = st.selectbox(
                        "학생 선택",
                        options=sorted_students,
                        format_func=lambda x: f"[{x[0]}번] {x[1]}"
                    )

                if selected_student:
                    target_num, target_name = selected_student
                    st.markdown(f"### {target_name} 학생의 기록 모음")

                    sub_tab_c, sub_tab_s, sub_tab_h = st.tabs(["창체 기록", "세특 기록", "행특 기록"])

                    with sub_tab_c:
                        if '창체' in st.session_state['data_store'] and st.session_state['data_store']['창체'] is not None:
                            c_item = st.session_state['data_store']['창체']
                            c_df, c_map = c_item['df'], c_item['col_map']
                            student_records = c_df[(get_safe_series(c_df, c_map['num_col']).astype(str).str.strip() == target_num) & (get_safe_series(c_df, c_map['name_col']).astype(str).str.strip() == target_name)]
                            if not student_records.empty:
                                for _, rec in student_records.iterrows():
                                    content_text = rec[c_map['content_col']]
                                    st.markdown(f"""
                                        <div class="student-card">
                                            <span class="subject-badge">창의적 체험활동</span>
                                            <span class="char-badge">{len(content_text)}자</span>
                                            <p style="margin-top:0.6rem; color:#334155; line-height:1.6;">{content_text}</p>
                                        </div>
                                    """, unsafe_allow_html=True)
                            else:
                                st.info("해당 학생의 창체 기록이 없습니다.")

                    with sub_tab_s:
                        if '세특' in st.session_state['data_store'] and st.session_state['data_store']['세특'] is not None:
                            s_item = st.session_state['data_store']['세특']
                            s_df, s_map = s_item['df'], s_item['col_map']
                            student_records = s_df[(get_safe_series(s_df, s_map['num_col']).astype(str).str.strip() == target_num) & (get_safe_series(s_df, s_map['name_col']).astype(str).str.strip() == target_name)]
                            if not student_records.empty:
                                for _, rec in student_records.iterrows():
                                    subj = rec.get('과목명', '기타')
                                    content_text = rec.get('내용', rec.get(s_map['content_col'], ''))
                                    char_cnt = rec.get('글자수', len(content_text))
                                    st.markdown(f"""
                                        <div class="student-card" style="border-left-color:#10B981;">
                                            <span class="subject-badge" style="background-color:#10B981;">{subj}</span>
                                            <span class="char-badge">{char_cnt}자</span>
                                            <p style="margin-top:0.6rem; color:#334155; line-height:1.6;">{content_text}</p>
                                        </div>
                                    """, unsafe_allow_html=True)
                            else:
                                st.info("해당 학생의 세특 기록이 없습니다.")

                    with sub_tab_h:
                        if '행특' in st.session_state['data_store'] and st.session_state['data_store']['행특'] is not None:
                            h_item = st.session_state['data_store']['행특']
                            h_df, h_map = h_item['df'], h_item['col_map']
                            student_records = h_df[(get_safe_series(h_df, h_map['num_col']).astype(str).str.strip() == target_num) & (get_safe_series(h_df, h_map['name_col']).astype(str).str.strip() == target_name)]
                            if not student_records.empty:
                                for _, rec in student_records.iterrows():
                                    content_text = rec[h_map['content_col']]
                                    st.markdown(f"""
                                        <div class="student-card" style="border-left-color:#8B5CF6;">
                                            <span class="subject-badge" style="background-color:#8B5CF6;">행동특성 및 종합의견</span>
                                            <span class="char-badge">{len(content_text)}자</span>
                                            <p style="margin-top:0.6rem; color:#334155; line-height:1.6;">{content_text}</p>
                                        </div>
                                    """, unsafe_allow_html=True)
                            else:
                                st.info("해당 학생의 행특 기록이 없습니다.")

    # --------------------------------------------------------------------------
    # MODE 3: 페이지 나눔 정제 검증 (관리 메뉴 모드)
    # --------------------------------------------------------------------------
    elif manage_mode == "페이지 나눔 정제 검증":
        st.subheader("지능형 페이지 나눔 정제 대조 검증")
        st.caption("페이지 나눔으로 인해 쪼개졌던 행들이 유실 없이 올바르게 병합되었는지 검증합니다.")

        inspect_type = st.radio("검증할 데이터 선택", ["세특", "창체", "행특"], horizontal=True)

        if inspect_type in st.session_state['data_store'] and st.session_state['data_store'][inspect_type] is not None:
            raw_df = st.session_state['data_store']['raw_data'][inspect_type]
            refined_data = st.session_state['data_store'][inspect_type]
            refined_df = refined_data['df']
            logs = st.session_state['data_store']['merge_logs'][inspect_type]

            col_m1, col_m2, col_m3, col_m4 = st.columns(4)
            with col_m1:
                st.markdown(f'<div class="metric-card"><div class="val">{len(raw_df)}</div><div class="lbl">원본 엑셀 행 수</div></div>', unsafe_allow_html=True)
            with col_m2:
                st.markdown(f'<div class="metric-card"><div class="val">{len(logs)}</div><div class="lbl">병합된 페이지 분할 행</div></div>', unsafe_allow_html=True)
            with col_m3:
                st.markdown(f'<div class="metric-card"><div class="val" style="color:#10B981;">{len(raw_df) - len(logs)}</div><div class="lbl">정제 후 실제 학생 수</div></div>', unsafe_allow_html=True)
            with col_m4:
                st.markdown(f'<div class="metric-card"><div class="val" style="color:#0EA5E9;">100.0%</div><div class="lbl">텍스트 데이터 보존율</div></div>', unsafe_allow_html=True)

            st.markdown("---")
            st.markdown("### 병합 처리된 행 상세 내역")
            if logs:
                log_df = pd.DataFrame(logs)
                log_df.columns = ["엑셀 행 번호", "대상 학생", "밀려 내려온 서술문", "결합 후 문장 끝 부분 요약"]
                st.dataframe(log_df, use_container_width=True)
            else:
                st.success("페이지 나눔으로 인해 밀려 내려온 행이 없습니다. (클린 데이터)")

            st.markdown("---")
            st.markdown("### 정제 완료 데이터 테이블 프리뷰 & 내보내기")
            display_df = clean_display_dataframe(refined_df)
            st.dataframe(display_df, use_container_width=True, height=380)

            col_m3_dl1, col_m3_dl2 = st.columns(2)
            with col_m3_dl1:
                st.download_button(
                    f"[{inspect_type}] 정제 원문 엑셀 저장 (.xlsx)",
                    data=create_formatted_excel_bytes({inspect_type: refined_df}),
                    file_name=f"{inspect_type}_정제원본.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"dl_m3_excel_{inspect_type}",
                    use_container_width=True
                )
            with col_m3_dl2:
                single_store = {
                    inspect_type: st.session_state['data_store'][inspect_type],
                    'file_names': st.session_state['data_store'].get('file_names', {})
                }
                st.download_button(
                    f"[{inspect_type}] 정제 원문 PDF 저장 (.pdf)",
                    data=create_refined_original_pdf_bytes(single_store),
                    file_name=f"{inspect_type}_정제원본.pdf",
                    mime="application/pdf",
                    key=f"dl_m3_pdf_{inspect_type}",
                    use_container_width=True
                )
        else:
            st.info(f"[{inspect_type}] 파일이 업로드되지 않았습니다.")

    # --------------------------------------------------------------------------
    # MODE 4: 데이터 분석 & 통계 (관리 메뉴 모드)
    # --------------------------------------------------------------------------
    elif manage_mode == "데이터 분석 & 통계":
        st.subheader("학생 기록 현황 분석 & 통계")
        inspect_type_stat = st.radio("분석할 데이터 선택", ["세특", "창체", "행특"], key="stat_radio", horizontal=True)

        if inspect_type_stat in st.session_state['data_store'] and st.session_state['data_store'][inspect_type_stat] is not None:
            data_item = st.session_state['data_store'][inspect_type_stat]
            df_stat = data_item['df']
            col_map = data_item['col_map']
            name_c = col_map['name_col']

            if inspect_type_stat == "세특" and '과목명' in df_stat.columns:
                col_chart1, col_chart2 = st.columns(2)
                with col_chart1:
                    st.markdown("#### 과목별 기록 분포")
                    subj_counts = df_stat['과목명'].value_counts()
                    st.bar_chart(subj_counts)
                with col_chart2:
                    st.markdown("#### 과목별 평균 글자 수")
                    avg_chars = df_stat.groupby('과목명')['글자수'].mean().round(1)
                    st.bar_chart(avg_chars)

                st.markdown("---")
                st.markdown("#### 학생별 세특 작성 과목 수 및 총 글자 수")
                student_summary = df_stat.groupby(name_c).agg(
                    과목수=('과목명', 'count'),
                    총글자수=('글자수', 'sum'),
                    평균글자수=('글자수', 'mean')
                ).reset_index()
                student_summary['평균글자수'] = student_summary['평균글자수'].round(1)
                st.dataframe(student_summary, use_container_width=True)
            else:
                content_c = col_map['content_col']
                st.markdown("#### 학생별 기록 글자 수 분포")
                df_stat['글자수'] = df_stat[content_c].astype(str).apply(len)
                char_chart_df = df_stat[[name_c, '글자수']].set_index(name_c)
                st.bar_chart(char_chart_df)
        else:
            st.info("분석할 데이터 파일이 업로드되지 않았습니다.")


if __name__ == "__main__":
    main()
